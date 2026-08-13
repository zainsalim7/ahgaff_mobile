"""
Smart Schedule Time Shifting (حلحلة الجدول) — backend tests
Endpoints under test:
  PUT    /api/weekly-schedule/{slot_id}   (duration_minutes -> cascading shift)
  POST   /api/weekly-schedule            (create -> resolve)
  DELETE /api/weekly-schedule/{slot_id}  (un-shift followers)
  GET    /api/weekly-schedule            (computed_start_time / computed_end_time)
  GET    /api/weekly-schedule/master-view
  GET    /api/weekly-schedule/export-visual/pdf|excel
  GET    /api/weekly-schedule/master-view/export/pdf|excel
  POST   /api/weekly-schedule/generate-lectures (dry_run, shift-aware)
"""
import io
import os

import pytest
import requests
from dotenv import dotenv_values

frontend_env = dotenv_values("/app/frontend/.env")
base_url = os.environ.get("REACT_APP_BACKEND_URL") or frontend_env.get("REACT_APP_BACKEND_URL")
if not base_url:
    raise RuntimeError("REACT_APP_BACKEND_URL missing from env and /app/frontend/.env")
BASE_URL = base_url.rstrip("/")

FACULTY_ID = "698e4f9297fef774e66e93a4"
DAY = "السبت"
# عقيدة level 5, slot 3 (11:30-13:00) -> extend to overflow into slot 4
AQEEDA_S3 = "6a527df95c8d6361dc30e0b3"
AQEEDA_S4 = "6a21981ee6cc2d7205c6c1bd"   # duration_minutes=75 (seeded, must stay)
T_S1 = "6a21981ee6cc2d7205c6c1b8"        # slot 1, teacher/room/section shared with T_S2
T_S2 = "6a21981ee6cc2d7205c6c1b9"
OTHER_S2 = "6a21981ee6cc2d7205c6c1ba"    # unrelated slot 2 lecture


@pytest.fixture(scope="module")
def client():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login", json={"username": "admin", "password": "admin123"}, timeout=60)
    if r.status_code != 200:
        pytest.fail(f"Admin login failed {r.status_code}: {r.text[:300]}")
    token = r.json().get("access_token")
    if not token:
        pytest.fail(f"No access_token in login response: {r.text[:300]}")
    s.headers.update({"Authorization": f"Bearer {token}"})
    return s


def slots_by_id(client):
    r = client.get(f"{BASE_URL}/api/weekly-schedule", params={"faculty_id": FACULTY_ID}, timeout=60)
    assert r.status_code == 200, r.text[:300]
    data = r.json()
    arr = data if isinstance(data, list) else data.get("slots", [])
    return {x["id"]: x for x in arr}


def set_duration(client, slot_id, minutes):
    r = client.put(f"{BASE_URL}/api/weekly-schedule/{slot_id}",
                   json={"duration_minutes": minutes}, timeout=90)
    return r


def pdf_text(content: bytes) -> str:
    from pypdf import PdfReader
    return "\n".join((p.extract_text() or "") for p in PdfReader(io.BytesIO(content)).pages)


def xlsx_cells(content: bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content))
    return [c.value for ws in wb.worksheets for row in ws.iter_rows() for c in row if isinstance(c.value, str)]


def has_range(text, a, b, seps=(" - ", "-", " – ")):
    return any((f"{a}{s}{b}" in text) or (f"{b}{s}{a}" in text) for s in seps)


# ---------- 0. baseline / serialization ----------
class TestSerialization:
    def test_list_exposes_computed_fields(self, client):
        m = slots_by_id(client)
        assert AQEEDA_S4 in m, "seeded عقيدة slot missing"
        for key in ("computed_start_time", "computed_end_time", "duration_minutes"):
            assert key in m[AQEEDA_S4], f"{key} missing from GET /api/weekly-schedule"
        assert m[AQEEDA_S4]["duration_minutes"] == 75
        assert m[AQEEDA_S4]["computed_end_time"] == "14:30"

    def test_master_view_exposes_computed_fields(self, client):
        r = client.get(f"{BASE_URL}/api/weekly-schedule/master-view",
                       params={"faculty_id": FACULTY_ID}, timeout=90)
        assert r.status_code == 200, r.text[:300]
        body = r.json()
        entries = []
        if isinstance(body, dict):
            for v in body.values():
                if isinstance(v, list):
                    entries += [e for e in v if isinstance(e, dict)]
        found = [e for e in entries if e.get("id") == AQEEDA_S4]
        assert found, f"seeded slot not present in master-view (keys={list(body)[:8]})"
        e = found[0]
        assert "computed_start_time" in e and "computed_end_time" in e
        assert e["computed_end_time"] == "14:30"


# ---------- 1. basic overflow shift + exports + generation ----------
class TestOverflowShift:
    def test_extend_slot3_shifts_slot4(self, client):
        r = set_duration(client, AQEEDA_S3, 120)
        assert r.status_code == 200, r.text[:400]
        body = r.json()
        assert "إزاحة تلقائية" in body.get("message", ""), body.get("message")
        shifted = body.get("shifted") or []
        assert isinstance(shifted, list) and shifted, "shifted array empty"
        target = [x for x in shifted if x.get("slot_id") == AQEEDA_S4]
        assert target, f"AQEEDA_S4 not in shifted: {shifted}"
        t = target[0]
        assert (t["from"], t["to"], t["end"]) == ("13:15", "13:30", "14:45"), t
        assert t.get("day") == DAY and t.get("slot_number") == 4

    def test_shift_persisted_in_list(self, client):
        m = slots_by_id(client)
        assert m[AQEEDA_S4]["computed_start_time"] == "13:30"
        assert m[AQEEDA_S4]["computed_end_time"] == "14:45"
        assert m[AQEEDA_S4]["duration_minutes"] == 75, "seeded duration must be untouched"
        assert m[AQEEDA_S3]["duration_minutes"] == 120
        assert m[AQEEDA_S3]["computed_end_time"] == "13:30"
        # unrelated slots untouched
        assert not m[OTHER_S2].get("computed_start_time")

    def test_visual_pdf_shows_shifted_time(self, client):
        r = client.get(f"{BASE_URL}/api/weekly-schedule/export-visual/pdf",
                       params={"faculty_id": FACULTY_ID}, timeout=180)
        assert r.status_code == 200, r.text[:300]
        assert r.content[:4] == b"%PDF"
        txt = pdf_text(r.content)
        assert has_range(txt, "13:30", "14:45"), "shifted range 13:30-14:45 not in visual PDF"

    def test_visual_excel_shows_shifted_time(self, client):
        r = client.get(f"{BASE_URL}/api/weekly-schedule/export-visual/excel",
                       params={"faculty_id": FACULTY_ID}, timeout=180)
        assert r.status_code == 200, r.text[:300]
        cells = xlsx_cells(r.content)
        assert any(has_range(c, "13:30", "14:45") for c in cells), \
            [c for c in cells if "⏱" in c][:6]

    def test_master_pdf_shows_shifted_time(self, client):
        r = client.get(f"{BASE_URL}/api/weekly-schedule/master-view/export/pdf",
                       params={"faculty_id": FACULTY_ID}, timeout=180)
        assert r.status_code == 200, r.text[:300]
        assert r.content[:4] == b"%PDF"
        txt = pdf_text(r.content)
        assert has_range(txt, "13:30", "14:45"), "shifted range not in master PDF"

    def test_master_excel_shows_shifted_time(self, client):
        r = client.get(f"{BASE_URL}/api/weekly-schedule/master-view/export/excel",
                       params={"faculty_id": FACULTY_ID}, timeout=180)
        assert r.status_code == 200, r.text[:300]
        cells = xlsx_cells(r.content)
        assert any(has_range(c, "13:30", "14:45") for c in cells), \
            [c for c in cells if "⏱" in c][:6]

    def test_generate_lectures_dry_run_with_active_shift(self, client):
        r = client.post(f"{BASE_URL}/api/weekly-schedule/generate-lectures",
                        json={"faculty_id": FACULTY_ID, "start_date": "2026-09-05",
                              "end_date": "2026-09-19", "dry_run": True}, timeout=180)
        assert r.status_code == 200, r.text[:500]
        body = r.json()
        assert isinstance(body, dict) and body, "empty dry-run response"
        blob = str(body)
        assert "13:30" in blob or "would_create" in blob or "to_create" in blob, blob[:400]
        print("DRY_RUN:", blob[:600])

    def test_clear_duration_restores_defaults(self, client):
        r = set_duration(client, AQEEDA_S3, 0)
        assert r.status_code == 200, r.text[:400]
        msg = r.json().get("message", "")
        assert "عادت الأوقات الافتراضية" in msg, msg
        assert not r.json().get("shifted")
        m = slots_by_id(client)
        assert not m[AQEEDA_S3].get("duration_minutes")
        assert not m[AQEEDA_S3].get("computed_start_time")
        assert not m[AQEEDA_S3].get("computed_end_time")
        assert not m[AQEEDA_S4].get("computed_start_time"), "follower start shift not cleared"
        # seeded slot keeps its own 75-min computed end
        assert m[AQEEDA_S4]["duration_minutes"] == 75
        assert m[AQEEDA_S4]["computed_end_time"] == "14:30"


# ---------- 2. cascading across resources ----------
class TestCascadingShift:
    def test_extend_slot1_shifts_only_related_slot2(self, client):
        r = set_duration(client, T_S1, 120)
        assert r.status_code == 200, r.text[:400]
        body = r.json()
        shifted = body.get("shifted") or []
        ids = {x["slot_id"] for x in shifted}
        assert T_S2 in ids, f"related slot-2 lecture not shifted: {shifted}"
        assert OTHER_S2 not in ids, "unrelated slot-2 lecture was shifted (should not be)"
        t = [x for x in shifted if x["slot_id"] == T_S2][0]
        assert (t["from"], t["to"]) == ("09:45", "10:00"), t
        assert t["end"] == "11:30", t
        assert "إزاحة تلقائية" in body.get("message", "")

    def test_cascade_persisted(self, client):
        m = slots_by_id(client)
        assert m[T_S1]["computed_end_time"] == "10:00"
        assert m[T_S2]["computed_start_time"] == "10:00"
        assert m[T_S2]["computed_end_time"] == "11:30"
        assert not m[OTHER_S2].get("computed_start_time")
        assert not m[OTHER_S2].get("computed_end_time")

    def test_reset_removes_all_computed(self, client):
        r = set_duration(client, T_S1, 0)
        assert r.status_code == 200, r.text[:400]
        m = slots_by_id(client)
        for sid in (T_S1, T_S2, OTHER_S2):
            assert not m[sid].get("computed_start_time"), sid
            assert not m[sid].get("computed_end_time"), sid
            assert not m[sid].get("duration_minutes"), sid


# ---------- 3. delete un-shifts followers (temp data on an empty day) ----------
class TestDeleteUnShift:
    created = []

    @pytest.fixture(scope="class", autouse=True)
    def cleanup(self, client):
        yield
        for sid in list(TestDeleteUnShift.created):
            client.delete(f"{BASE_URL}/api/weekly-schedule/{sid}", timeout=60)

    def test_delete_extended_lecture_unshifts_follower(self, client):
        src = slots_by_id(client)[T_S1]
        base = {
            "faculty_id": FACULTY_ID,
            "department_id": src["department_id"],
            "level": src["level"],
            "section": "TEST_Z",
            "day": "الخميس",
            "course_id": src["course_id"],
            "teacher_id": src["teacher_id"],
            "room_id": src["room_id"],
        }
        r1 = client.post(f"{BASE_URL}/api/weekly-schedule", json={**base, "slot_number": 1,
                                                                  "duration_minutes": 120}, timeout=60)
        assert r1.status_code in (200, 201), r1.text[:400]
        id1 = r1.json().get("id")
        TestDeleteUnShift.created.append(id1)

        r2 = client.post(f"{BASE_URL}/api/weekly-schedule", json={**base, "slot_number": 2}, timeout=60)
        assert r2.status_code in (200, 201), r2.text[:400]
        id2 = r2.json().get("id")
        TestDeleteUnShift.created.append(id2)
        shifted_ids = {x["slot_id"] for x in (r2.json().get("shifted") or [])}
        assert id2 in shifted_ids, f"newly created follower not shifted: {r2.json()}"

        m = slots_by_id(client)
        assert m[id2]["computed_start_time"] == "10:00", m[id2]
        assert m[id2]["computed_end_time"] == "11:30", m[id2]

        d = client.delete(f"{BASE_URL}/api/weekly-schedule/{id1}", timeout=90)
        assert d.status_code == 200, d.text[:300]
        TestDeleteUnShift.created.remove(id1)

        m = slots_by_id(client)
        assert not m[id2].get("computed_start_time"), "follower still shifted after deleting extended lecture"
        assert not m[id2].get("computed_end_time")


# ---------- 4. regression: normal updates & conflicts ----------
class TestRegression:
    def test_room_change_and_revert(self, client):
        before = slots_by_id(client)[OTHER_S2]
        old_room = before["room_id"]
        new_room = "69e8cf9804f85810f65e6348"
        assert old_room != new_room
        r = client.put(f"{BASE_URL}/api/weekly-schedule/{OTHER_S2}",
                       json={"room_id": new_room}, timeout=90)
        assert r.status_code == 200, r.text[:400]
        assert r.json().get("message")
        assert slots_by_id(client)[OTHER_S2]["room_id"] == new_room
        rb = client.put(f"{BASE_URL}/api/weekly-schedule/{OTHER_S2}",
                        json={"room_id": old_room}, timeout=90)
        assert rb.status_code == 200, rb.text[:400]
        assert slots_by_id(client)[OTHER_S2]["room_id"] == old_room

    def test_room_conflict_returns_409(self, client):
        m = slots_by_id(client)
        # OTHER_S2 (slot 2) taking T_S2's room (same day+slot) must conflict
        r = client.put(f"{BASE_URL}/api/weekly-schedule/{OTHER_S2}",
                       json={"room_id": m[T_S2]["room_id"]}, timeout=90)
        assert r.status_code == 409, f"expected 409, got {r.status_code}: {r.text[:300]}"
        assert slots_by_id(client)[OTHER_S2]["room_id"] == m[OTHER_S2]["room_id"], "state changed on conflict"

    def test_invalid_slot_id_404(self, client):
        r = client.put(f"{BASE_URL}/api/weekly-schedule/6a21981ee6cc2d7205c6c000",
                       json={"duration_minutes": 100}, timeout=60)
        assert r.status_code == 404, r.status_code

    def test_final_state_intact(self, client):
        m = slots_by_id(client)
        assert m[AQEEDA_S4]["duration_minutes"] == 75
        assert m[AQEEDA_S4]["computed_end_time"] == "14:30"
        assert not m[AQEEDA_S3].get("duration_minutes")
        assert not m[T_S1].get("duration_minutes")
