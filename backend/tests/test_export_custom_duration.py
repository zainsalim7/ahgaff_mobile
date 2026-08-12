"""
Tests for custom lecture duration (duration_minutes) appearing in exported PDF/Excel
Endpoints under test:
  GET /api/weekly-schedule/export-visual/pdf
  GET /api/weekly-schedule/export-visual/excel
  GET /api/weekly-schedule/master-view/export/pdf
  GET /api/weekly-schedule/master-view/export/excel
  PUT /api/weekly-schedule/{slot_id}  (duration_minutes set/clear)
"""
import io
import os
import re

import pytest
import requests
from dotenv import dotenv_values

frontend_env = dotenv_values("/app/frontend/.env")
base_url = os.environ.get("REACT_APP_BACKEND_URL") or frontend_env.get("REACT_APP_BACKEND_URL")
if not base_url:
    raise RuntimeError("REACT_APP_BACKEND_URL missing")
BASE_URL = base_url.rstrip("/")

FACULTY_ID = "698e4f9297fef774e66e93a4"
SEEDED_SLOT_ID = "6a21981ee6cc2d7205c6c1bd"
EXPECTED_START = "13:15"
EXPECTED_END = "14:30"


@pytest.fixture(scope="module")
def client():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login", json={"username": "admin", "password": "admin123"}, timeout=60)
    if r.status_code != 200:
        pytest.fail(f"Admin login failed {r.status_code}: {r.text[:300]}")
    token = r.json().get("access_token")
    if not token:
        pytest.fail(f"No access_token in login response: {r.text[:300]}")
    s.headers.update({"Authorization": f"Bearer {token}"})
    return s


def pdf_text(content: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(content))
    return "\n".join((p.extract_text() or "") for p in reader.pages)


def xlsx_cells(content: bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content))
    vals = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str):
                    vals.append(c.value)
    return vals


# ---- helper: time math sanity on the backend helper itself ----
def test_add_minutes_helper():
    import sys
    for p in ("/app/backend", "/app/backend/backend"):
        if p not in sys.path:
            sys.path.insert(0, p)
    from backend.routes.weekly_schedule import _add_minutes
    assert _add_minutes("13:15", 75) == "14:30"
    assert _add_minutes("08:00", 90) == "09:30"
    assert _add_minutes("23:30", 45) == "00:15"
    assert _add_minutes("bad", 45) == "bad"


class TestVisualExports:
    def test_export_visual_pdf(self, client):
        r = client.get(f"{BASE_URL}/api/weekly-schedule/export-visual/pdf",
                       params={"faculty_id": FACULTY_ID}, timeout=120)
        assert r.status_code == 200, r.text[:400]
        assert r.content[:4] == b"%PDF", "response is not a valid PDF"
        assert len(r.content) > 2000
        txt = pdf_text(r.content)
        # Arabic RTL extraction may reverse order -> accept either direction
        assert (f"{EXPECTED_START} - {EXPECTED_END}" in txt) or (f"{EXPECTED_END} - {EXPECTED_START}" in txt), \
            f"computed time range not found in PDF text; times found: {re.findall(r'[0-9]{2}:[0-9]{2}', txt)[:20]}"

    def test_export_visual_excel(self, client):
        r = client.get(f"{BASE_URL}/api/weekly-schedule/export-visual/excel",
                       params={"faculty_id": FACULTY_ID}, timeout=120)
        assert r.status_code == 200, r.text[:400]
        assert r.content[:2] == b"PK", "response is not a valid XLSX"
        cells = xlsx_cells(r.content)
        marked = [c for c in cells if "⏱" in c]
        assert marked, "no cell with ⏱ duration marker found"
        assert any(f"⏱ {EXPECTED_START} - {EXPECTED_END}" in c for c in marked), marked[:5]
        # regression: only the slot(s) with duration_minutes carry the marker
        assert len(marked) == 1, f"expected exactly 1 cell with time marker, got {len(marked)}: {marked}"


class TestMasterExports:
    def test_master_pdf(self, client):
        r = client.get(f"{BASE_URL}/api/weekly-schedule/master-view/export/pdf",
                       params={"faculty_id": FACULTY_ID}, timeout=120)
        assert r.status_code == 200, r.text[:400]
        assert r.content[:4] == b"%PDF"
        txt = pdf_text(r.content)
        assert (f"{EXPECTED_START}-{EXPECTED_END}" in txt) or (f"{EXPECTED_END}-{EXPECTED_START}" in txt), \
            f"computed time not in master PDF; times: {re.findall(r'[0-9]{2}:[0-9]{2}', txt)[:20]}"

    def test_master_excel(self, client):
        r = client.get(f"{BASE_URL}/api/weekly-schedule/master-view/export/excel",
                       params={"faculty_id": FACULTY_ID}, timeout=120)
        assert r.status_code == 200, r.text[:400]
        assert r.content[:2] == b"PK"
        cells = xlsx_cells(r.content)
        marked = [c for c in cells if "⏱" in c]
        assert any(f"⏱ {EXPECTED_START} - {EXPECTED_END}" in c for c in marked), marked[:5]
        assert len(marked) == 1, f"expected exactly 1 marked cell, got {len(marked)}: {marked}"

    def test_master_excel_missing_faculty_id(self, client):
        r = client.get(f"{BASE_URL}/api/weekly-schedule/master-view/export/excel", timeout=60)
        assert r.status_code == 422, r.status_code


class TestDurationUpdateRoundTrip:
    """PUT duration_minutes on a second slot, re-export, then clear it."""

    @staticmethod
    def master_entries(client):
        r = client.get(f"{BASE_URL}/api/weekly-schedule/master-view",
                       params={"faculty_id": FACULTY_ID}, timeout=60)
        assert r.status_code == 200, r.text[:300]
        return r.json().get("entries", [])

    @pytest.fixture(scope="class")
    def other_slot(self, client):
        cand = [s for s in self.master_entries(client)
                if s.get("id") != SEEDED_SLOT_ID and not s.get("duration_minutes")]
        if not cand:
            pytest.skip("no other slot without duration_minutes available")
        return cand[0]

    def test_set_and_clear_duration(self, client, other_slot):
        slot_id = other_slot["id"]
        try:
            r = client.put(f"{BASE_URL}/api/weekly-schedule/{slot_id}",
                           json={"duration_minutes": 45}, timeout=60)
            assert r.status_code == 200, r.text[:400]

            # verify persisted
            got = [s for s in self.master_entries(client) if s.get("id") == slot_id]
            assert got and got[0].get("duration_minutes") == 45, got[:1]

            # export should now show 2 marked cells
            e = client.get(f"{BASE_URL}/api/weekly-schedule/export-visual/excel",
                           params={"faculty_id": FACULTY_ID}, timeout=120)
            assert e.status_code == 200
            marked = [c for c in xlsx_cells(e.content) if "⏱" in c]
            assert len(marked) == 2, f"expected 2 marked cells after update, got {len(marked)}: {marked}"
        finally:
            c = client.put(f"{BASE_URL}/api/weekly-schedule/{slot_id}",
                           json={"duration_minutes": 0}, timeout=60)
            assert c.status_code == 200, c.text[:300]

        # after clearing, only the seeded slot remains marked
        e2 = client.get(f"{BASE_URL}/api/weekly-schedule/export-visual/excel",
                        params={"faculty_id": FACULTY_ID}, timeout=120)
        marked2 = [c for c in xlsx_cells(e2.content) if "⏱" in c]
        assert len(marked2) == 1, f"clearing duration_minutes=0 did not remove marker: {marked2}"

    def test_seeded_slot_intact(self, client):
        seeded = [s for s in self.master_entries(client) if s.get("id") == SEEDED_SLOT_ID]
        assert seeded and seeded[0].get("duration_minutes") == 75, "seeded 75-min slot was altered"

    def test_list_endpoint_exposes_duration(self, client):
        """GET /api/weekly-schedule should expose duration_minutes for round-trip editing."""
        r = client.get(f"{BASE_URL}/api/weekly-schedule", params={"faculty_id": FACULTY_ID}, timeout=60)
        assert r.status_code == 200
        data = r.json()
        slots = data if isinstance(data, list) else data.get("slots", [])
        seeded = [s for s in slots if s.get("id") == SEEDED_SLOT_ID]
        assert seeded, "seeded slot missing from list endpoint"
        assert seeded[0].get("duration_minutes") == 75, \
            f"duration_minutes not returned by list endpoint: keys={sorted(seeded[0].keys())}"


class TestFilteredExports:
    """Exports with narrower filters must still render the computed time and not crash."""

    def test_visual_excel_filtered_by_dept_level(self, client):
        r = client.get(f"{BASE_URL}/api/weekly-schedule/export-visual/excel",
                       params={"faculty_id": FACULTY_ID,
                               "department_id": "698f8093539792f8917b7bd3", "level": 5}, timeout=120)
        assert r.status_code == 200, r.text[:400]
        marked = [c for c in xlsx_cells(r.content) if "⏱" in c]
        assert any(f"⏱ {EXPECTED_START} - {EXPECTED_END}" in c for c in marked), marked[:5]

    def test_visual_pdf_filtered_by_dept_level(self, client):
        r = client.get(f"{BASE_URL}/api/weekly-schedule/export-visual/pdf",
                       params={"faculty_id": FACULTY_ID,
                               "department_id": "698f8093539792f8917b7bd3", "level": 5}, timeout=120)
        assert r.status_code == 200, r.text[:400]
        assert r.content[:4] == b"%PDF"
        txt = pdf_text(r.content)
        assert EXPECTED_END in txt

    def test_master_pdf_filtered_by_dept(self, client):
        r = client.get(f"{BASE_URL}/api/weekly-schedule/master-view/export/pdf",
                       params={"faculty_id": FACULTY_ID, "department_id": "698f8093539792f8917b7bd3"}, timeout=120)
        assert r.status_code == 200, r.text[:400]
        assert r.content[:4] == b"%PDF"

    def test_master_pdf_unknown_faculty_no_crash(self, client):
        r = client.get(f"{BASE_URL}/api/weekly-schedule/master-view/export/pdf",
                       params={"faculty_id": "6a00000000000000000000ff"}, timeout=120)
        assert r.status_code in (200, 400, 404), f"{r.status_code}: {r.text[:300]}"

    def test_exports_require_auth(self):
        r = requests.get(f"{BASE_URL}/api/weekly-schedule/export-visual/pdf",
                         params={"faculty_id": FACULTY_ID}, timeout=60)
        assert r.status_code in (401, 403), r.status_code
