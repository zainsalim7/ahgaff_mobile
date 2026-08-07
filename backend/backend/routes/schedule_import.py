"""
استيراد الجدول الأسبوعي الشامل من Excel + توليد قالب الاستيراد
"""
import io
import re
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from bson import ObjectId
from pymongo.errors import DuplicateKeyError

from .deps import get_db, get_current_user, log_activity
from .weekly_schedule import can_manage_schedule, _is_period_unavailable, _build_master_data, _sync_future_lectures

router = APIRouter(tags=["استيراد الجدول الأسبوعي"])

AR_ORDINALS = {
    "الاول": 1, "الأول": 1, "الثاني": 2, "الثالث": 3, "الرابع": 4, "الخامس": 5,
    "السادس": 6, "السابع": 7, "الثامن": 8, "التاسع": 9, "العاشر": 10,
}
KNOWN_DAYS = {"السبت", "الأحد", "الاثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة"}
SECTION_CHARS = {"أ", "ا", "ب", "ج", "د", "هـ", "ه", "و", "ز", "ح"}


def _norm(s) -> str:
    """تطبيع عربي للمطابقة النصية: همزات/تاء مربوطة/ياء + مسافات"""
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r"[\u064B-\u0652\u0640]", "", s)  # تشكيل + تطويل
    s = s.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    s = s.replace("ة", "ه").replace("ى", "ي")
    s = re.sub(r"\s+", " ", s)
    return s


def _norm_day(s) -> str:
    n = _norm(s)
    for d in KNOWN_DAYS:
        if _norm(d) == n:
            return d
    return ""


def _parse_group_label(label: str):
    """يستخرج (المستوى، الشعبة) من تسمية المجموعة مثل: 'المستوى 2 - شعبة أ' أو 'الثاني شريعة أ'"""
    txt = _norm(label)
    if not txt:
        return None, None
    level = None
    m = re.search(r"\d+", txt)
    if m:
        level = int(m.group())
    else:
        for word, val in AR_ORDINALS.items():
            if _norm(word) in txt.split():
                level = val
                break
    section = ""
    m = re.search(r"شعبه\s+(\S+)", txt)
    if m:
        section = m.group(1)
    else:
        last = txt.split()[-1] if txt.split() else ""
        if last in {_norm(c) for c in SECTION_CHARS} and len(txt.split()) > 1:
            section = last
    return level, section


ROW_KINDS = [("course", "مقرر"), ("room", "قاع"), ("teacher", "استاذ"), ("teacher", "معلم")]


def _row_kind(b_label: str):
    n = _norm(b_label)
    for kind, key in ROW_KINDS:
        if key in n:
            return kind
    return None


@router.get("/weekly-schedule/import-template")
async def download_import_template(
    faculty_id: str,
    department_id: str,
    current_user: dict = Depends(get_current_user),
):
    """قالب Excel للاستيراد بنفس بنية النموذج: مجموعات × (أيام × فترات)، كل خلية 3 صفوف (مقرر/قاعة/أستاذ)"""
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db = get_db()
    data = await _build_master_data(db, faculty_id, department_id)
    working_days = data["working_days"]
    time_slots = sorted(data["time_slots"], key=lambda x: x.get("slot_number", 0))
    groups = [g for g in data["groups"] if g["department_id"] == department_id]
    if not working_days or not time_slots:
        raise HTTPException(status_code=400, detail="لا توجد أيام عمل أو فترات معرفة لهذه الكلية — عرّفها من الإعدادات أولاً")

    # 🆕 القالب متاح لكل الأقسام: المجموعات تُبنى من كل مقررات القسم حتى غير المسندة لأستاذ
    all_courses = await db.courses.find({"department_id": department_id, "is_active": True}).to_list(2000)
    seen_groups = {(g["level"], g["section"]) for g in groups}
    for cdoc in all_courses:
        key = (cdoc.get("level") or 1, cdoc.get("section") or "")
        if key not in seen_groups:
            seen_groups.add(key)
            groups.append({"level": key[0], "section": key[1], "department_id": department_id})
    groups.sort(key=lambda g: (g["level"], g["section"]))
    if not groups:
        raise HTTPException(status_code=400, detail="لا توجد مقررات في هذا القسم — أضف المقررات أولاً ثم حمّل القالب")

    dept = await db.departments.find_one({"_id": ObjectId(department_id)})
    faculty = await db.faculties.find_one({"_id": ObjectId(faculty_id)})

    # خريطة الخلايا الموجودة (لتعبئة القالب بالجدول الحالي)
    cell_map = {}
    for e in data["entries"]:
        if e["department_id"] != department_id:
            continue
        cell_map[(e["level"], e["section"], e["day"], e["slot_number"])] = e

    wb = Workbook()
    ws = wb.active
    ws.title = "الجدول"
    ws.sheet_view.rightToLeft = True

    ns = len(time_slots)
    ncols = 2 + len(working_days) * ns

    thin = Side(style="thin", color="B8C4D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tc = ws.cell(row=1, column=1, value=f"قالب استيراد الجدول الأسبوعي — {(faculty or {}).get('name', '')} — قسم {(dept or {}).get('name', '')}")
    tc.font = Font(bold=True, size=13, color="0D2A52")
    tc.alignment = center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ic = ws.cell(row=2, column=1, value="⚠️ استخدم الأسماء كما في ورقة (الأدلة) حرفياً. لا تغيّر تسميات المستويات أو الصفوف. الخلايا الفارغة تُتجاهل.")
    ic.font = Font(size=9, color="B26A00", bold=True)
    ic.alignment = Alignment(horizontal="right", vertical="center")

    # صف الأيام (3) + صف الفترات (4)
    for h, col in (("المستوى / الشعبة", 1), ("", 2)):
        c = ws.cell(row=3, column=col, value=h)
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
        c.fill = PatternFill("solid", fgColor="0D2A52")
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = center
    for di, day in enumerate(working_days):
        c1 = 3 + di * ns
        ws.merge_cells(start_row=3, start_column=c1, end_row=3, end_column=c1 + ns - 1)
        dc = ws.cell(row=3, column=c1, value=day)
        dc.fill = PatternFill("solid", fgColor="1565C0")
        dc.font = Font(bold=True, color="FFFFFF", size=11)
        dc.alignment = center
        for si, ts in enumerate(time_slots):
            sc = ws.cell(row=4, column=c1 + si, value=ts.get("name") or f"الفترة {ts.get('slot_number')}")
            sc.fill = PatternFill("solid", fgColor="3D7EDE")
            sc.font = Font(bold=True, color="FFFFFF", size=8)
            sc.alignment = center
            sc.border = border

    row_labels = ["المقرر", "القاعة", "الأستاذ"]
    r = 5
    for g in groups:
        label = f"المستوى {g['level']}" + (f" - شعبة {g['section']}" if g["section"] else "")
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=1)
        gc = ws.cell(row=r, column=1, value=label)
        gc.fill = PatternFill("solid", fgColor="EEF3FA")
        gc.font = Font(bold=True, size=9, color="1A2540")
        gc.alignment = center
        gc.border = border
        for i, rl in enumerate(row_labels):
            bc = ws.cell(row=r + i, column=2, value=rl)
            bc.fill = PatternFill("solid", fgColor="F5F7FA")
            bc.font = Font(size=8, color="5A6B85", bold=True)
            bc.alignment = center
            bc.border = border
            for di, day in enumerate(working_days):
                for si, ts in enumerate(time_slots):
                    col = 3 + di * ns + si
                    cell = ws.cell(row=r + i, column=col)
                    cell.border = border
                    cell.alignment = center
                    cell.font = Font(size=8)
                    e = cell_map.get((g["level"], g["section"], day, ts.get("slot_number")))
                    if e:
                        cell.value = [e["course_name"], e["room_name"], e["teacher_name"]][i]
                        cell.fill = PatternFill("solid", fgColor="FFF8E1")
        r += 3

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 9
    for c in range(3, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # ورقة الأدلة: كل مقررات القسم (حتى غير المسندة) + القاعات + كل أساتذة القسم
    ws2 = wb.create_sheet("الأدلة")
    ws2.sheet_view.rightToLeft = True
    courses = all_courses
    t_ids = [c.get("teacher_id") for c in courses if c.get("teacher_id")]
    teachers = {str(t["_id"]): t.get("full_name", "") for t in await db.teachers.find({"_id": {"$in": [ObjectId(x) for x in t_ids]}}).to_list(1000)} if t_ids else {}
    rooms = await db.rooms.find({"faculty_id": faculty_id, "is_active": True}).to_list(300)
    dept_teachers = await db.teachers.find({"department_id": department_id, "is_active": True}).to_list(1000)
    heads = ["اسم المقرر (انسخه حرفياً)", "المستوى", "الشعبة", "الأستاذ المسند", "", "القاعات المتاحة", "", "أساتذة القسم (استخدم أي اسم للإسناد)"]
    for ci, h in enumerate(heads, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="1565C0")
        c.alignment = center
    for ri, cdoc in enumerate(sorted(courses, key=lambda x: (x.get("level") or 0, x.get("section") or "", x.get("name") or "")), 2):
        ws2.cell(row=ri, column=1, value=cdoc.get("name", ""))
        ws2.cell(row=ri, column=2, value=cdoc.get("level") or 1)
        ws2.cell(row=ri, column=3, value=cdoc.get("section") or "-")
        ws2.cell(row=ri, column=4, value=teachers.get(cdoc.get("teacher_id", ""), "⚠️ بلا أستاذ — اكتب اسم الأستاذ في الجدول وسيُسند تلقائياً"))
    for ri, rdoc in enumerate(sorted(rooms, key=lambda x: x.get("name", "")), 2):
        ws2.cell(row=ri, column=6, value=rdoc.get("name", ""))
    for ri, tdoc in enumerate(sorted(dept_teachers, key=lambda x: x.get("full_name", "")), 2):
        ws2.cell(row=ri, column=8, value=tdoc.get("full_name", ""))
    for ci, w in enumerate([35, 10, 10, 40, 4, 22, 4, 30], 1):
        from openpyxl.utils import get_column_letter as gcl
        ws2.column_dimensions[gcl(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"schedule_import_template_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/weekly-schedule/import-master")
async def import_master_schedule(
    faculty_id: str = Form(...),
    department_id: str = Form(...),
    dry_run: str = Form("1"),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """استيراد الجدول الأسبوعي الشامل من Excel لقسم محدد.
    السياسة: الإكسل هو الأساس — الخلايا المعبأة في الملف تستبدل ما يقابلها في النظام (استبدال على مستوى الخلية)
    + الخلايا الفارغة في الملف لا تمس الموجود + أخطاء الأسماء تُتخطى مع التقرير + أي تعارض جدولة يوقف الاستيراد كاملاً.
    """
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    is_dry = str(dry_run).strip() not in ("0", "false", "False")

    from openpyxl import load_workbook
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="تعذر قراءة الملف — تأكد أنه ملف Excel (.xlsx) سليم")
    ws = wb.worksheets[0]

    db = get_db()
    dept = await db.departments.find_one({"_id": ObjectId(department_id)})
    if not dept or dept.get("faculty_id") != faculty_id:
        raise HTTPException(status_code=400, detail="القسم غير موجود أو لا يتبع الكلية المحددة")

    settings = await db.schedule_settings.find_one({"_id": f"faculty_{faculty_id}"}) or await db.schedule_settings.find_one({"_id": "global"})
    time_slots = sorted((settings or {}).get("time_slots", []), key=lambda x: x.get("slot_number", 0))
    working_days = (settings or {}).get("working_days", [])
    if not time_slots or not working_days:
        raise HTTPException(status_code=400, detail="لا توجد فترات أو أيام عمل معرفة لهذه الكلية")

    # ===== 1) تحديد صف الأيام وصف الفترات =====
    days_row = None
    for r in range(1, min(11, ws.max_row + 1)):
        found = sum(1 for c in range(1, ws.max_column + 1) if _norm_day(ws.cell(row=r, column=c).value))
        if found >= 2:
            days_row = r
            break
    if not days_row:
        raise HTTPException(status_code=400, detail="لم يتم العثور على صف الأيام (السبت، الأحد...) في الملف — استخدم القالب الرسمي")
    periods_row = days_row + 1

    # خريطة عمود → (يوم، ترتيب الفترة داخل اليوم)
    col_day = {}
    current_day, order = "", 0
    for c in range(1, ws.max_column + 1):
        d = _norm_day(ws.cell(row=days_row, column=c).value)
        if d:
            current_day, order = d, 0
        if current_day and ws.cell(row=periods_row, column=c).value not in (None, ""):
            order += 1
            col_day[c] = (current_day, order)

    errors, conflicts, skipped_existing = [], [], []
    slot_by_order = {i + 1: ts for i, ts in enumerate(time_slots)}
    working_set = set(working_days)

    # ===== 2) بيانات النظام للمطابقة =====
    courses = await db.courses.find({"department_id": department_id, "is_active": True}).to_list(2000)
    courses_by_name = {}
    for cdoc in courses:
        courses_by_name.setdefault(_norm(cdoc.get("name")), []).append(cdoc)
    courses_by_id = {str(c["_id"]): c.get("name", "") for c in courses}
    all_teachers = await db.teachers.find({}).to_list(3000)
    teachers_map = {str(t["_id"]): t for t in all_teachers}
    teachers_by_name = {}
    for t in all_teachers:
        teachers_by_name.setdefault(_norm(t.get("full_name")), []).append(t)
    rooms = await db.rooms.find({"faculty_id": faculty_id, "is_active": True}).to_list(300)
    rooms_by_name = {_norm(r.get("name")): r for r in rooms}
    rooms_by_id = {str(r["_id"]): r.get("name", "") for r in rooms}

    existing_slots = await db.weekly_schedule.find({"department_id": department_id}).to_list(5000)
    existing_cells = {(s.get("level"), s.get("section", "") or "", s.get("day"), s.get("slot_number")): s for s in existing_slots}

    # ===== 3) المرور على كتل المجموعات (3 صفوف لكل مجموعة) =====
    to_create = []
    reassign_map = {}  # course_id -> {"new_id", "new_name", "old_id", "old_name", "course_name"}
    file_teachers = {}  # course_id -> {normalized_name: raw_name} لكشف التناقضات
    file_positions = {}  # course_id -> {(day, slot_number)} مواضع المقرر كما في الملف (الملف هو الأساس حرفياً)
    r = periods_row + 1
    while r <= ws.max_row:
        label = ws.cell(row=r, column=1).value
        if label is None or not str(label).strip():
            r += 1
            continue
        level, section = _parse_group_label(str(label))
        block_rows = {"course": r, "room": r + 1, "teacher": r + 2}
        for i in range(3):
            kind = _row_kind(str(ws.cell(row=r + i, column=2).value or ""))
            if kind:
                block_rows[kind] = r + i
        if level is None:
            errors.append(f"صف {r}: تعذر تحديد المستوى من التسمية '{str(label).strip()}' — تُخُطيت المجموعة")
            r += 3
            continue

        for c, (day, order) in col_day.items():
            course_txt = str(ws.cell(row=block_rows["course"], column=c).value or "").strip()
            room_txt = str(ws.cell(row=block_rows["room"], column=c).value or "").strip()
            teacher_txt = str(ws.cell(row=block_rows["teacher"], column=c).value or "").strip()
            if not course_txt and not room_txt and not teacher_txt:
                continue
            loc = f"[المستوى {level}{' شعبة ' + section if section else ''} — {day} فترة {order}]"
            if day not in working_set:
                errors.append(f"{loc} اليوم '{day}' ليس من أيام عمل الكلية — تُخُطيت الخلية")
                continue
            ts = slot_by_order.get(order)
            if not ts:
                errors.append(f"{loc} الفترة رقم {order} غير معرفة في إعدادات الكلية ({len(time_slots)} فترات فقط) — تُخُطيت الخلية")
                continue
            slot_number = ts.get("slot_number")
            if not course_txt:
                errors.append(f"{loc} خلية بلا اسم مقرر — تُخُطيت")
                continue

            # مطابقة المقرر (مقارنة مستوى/شعبة مطبّعة + تمييز الأسماء المتطابقة بالأستاذ)
            candidates = courses_by_name.get(_norm(course_txt), [])
            nsec = _norm(section)
            matches = [x for x in candidates if (x.get("level") or 1) == level and _norm(x.get("section") or "") == nsec]
            if not matches:
                # 🔗 مقرر مشترك: مستوى الخلية ضمن المستويات المشتركة للمقرر
                matches = [x for x in candidates if level in (x.get("shared_levels") or []) and _norm(x.get("section") or "") == nsec]
            cross_level = False
            if not matches:
                # 🔗 محاضرة مشتركة عبر المستويات: قبول مقرر من مستوى آخر بشرط الدمج (يُتحقق لاحقاً)
                cl = [x for x in candidates if _norm(x.get("section") or "") == nsec]
                if len(cl) > 1 and teacher_txt:
                    tcl = [x for x in cl if _norm(teachers_map.get(x.get("teacher_id", "") or "", {}).get("full_name", "")) == _norm(teacher_txt)]
                    if tcl:
                        cl = tcl
                if len(cl) == 1:
                    matches = cl
                    cross_level = True
            if len(matches) > 1 and teacher_txt:
                tmatches = [x for x in matches if _norm(teachers_map.get(x.get("teacher_id", ""), {}).get("full_name", "")) == _norm(teacher_txt)]
                if tmatches:
                    matches = tmatches
            course = matches[0] if matches else None
            if not course:
                if candidates:
                    have = "، ".join(f"م{x.get('level') or 1}{'/' + x.get('section') if x.get('section') else ''}" for x in candidates[:3])
                    errors.append(f"{loc} المقرر '{course_txt}' موجود لكن لمستوى/شعبة مختلفة ({have}) — تُخُطيت الخلية")
                else:
                    errors.append(f"{loc} المقرر '{course_txt}' غير موجود في القسم — تُخُطيت الخلية")
                continue
            section_val = course.get("section") or ""
            cid_str = str(course["_id"])
            file_positions.setdefault(cid_str, set()).add((day, slot_number))
            assigned = teachers_map.get(course.get("teacher_id", "") or "", {})
            assigned_norm = _norm(assigned.get("full_name", ""))

            # 🧑‍🏫 الإكسل هو الأساس في الإسناد أيضاً:
            # - مقرر بلا إسناد + اسم أستاذ في الملف → يُسند إليه
            # - اسم أستاذ مختلف عن المسند → يُستبدل الإسناد بما في الملف
            if teacher_txt and _norm(teacher_txt) != assigned_norm:
                cands = teachers_by_name.get(_norm(teacher_txt), [])
                if not cands:
                    errors.append(f"{loc} الأستاذ '{teacher_txt}' غير موجود في النظام — تُخُطيت الخلية")
                    continue
                if len(cands) > 1:
                    errors.append(f"{loc} يوجد أكثر من معلم بالاسم '{teacher_txt}' في النظام — تُخُطيت الخلية")
                    continue
                new_tid = str(cands[0]["_id"])
                prev = reassign_map.get(cid_str)
                if prev and prev["new_id"] != new_tid:
                    conflicts.append(f"{loc} تناقض داخل الملف: المقرر '{course.get('name', '')}' مذكور بأستاذين مختلفين ('{prev['new_name']}' و'{teacher_txt}')")
                    continue
                if not prev:
                    reassign_map[cid_str] = {
                        "new_id": new_tid, "new_name": cands[0].get("full_name", ""),
                        "old_id": course.get("teacher_id") or "", "old_name": assigned.get("full_name", ""),
                        "course_name": course.get("name", ""),
                    }
            elif not course.get("teacher_id") and cid_str not in reassign_map:
                errors.append(f"{loc} المقرر '{course_txt}' بلا أستاذ مسند في النظام ولا اسم أستاذ في الملف — تُخُطيت الخلية")
                continue

            effective_tid = reassign_map[cid_str]["new_id"] if cid_str in reassign_map else course["teacher_id"]
            teacher = teachers_map.get(effective_tid, {})
            if teacher_txt:
                file_teachers.setdefault(cid_str, {})[_norm(teacher_txt)] = teacher_txt
            if not room_txt:
                errors.append(f"{loc} القاعة مطلوبة — تُخُطيت الخلية")
                continue
            room = rooms_by_name.get(_norm(room_txt))
            if not room:
                errors.append(f"{loc} القاعة '{room_txt}' غير مسجلة في الكلية — تُخُطيت الخلية")
                continue

            # 🔁 الخلية مشغولة مسبقاً → الإكسل هو الأساس: استبدال (أو تخطٍ إن كانت مطابقة تماماً)
            replace_id = None
            replace_desc = ""
            replace_teacher = ""
            replace_course = ""
            ex = existing_cells.get((level, section_val, day, slot_number))
            if ex:
                same = ex.get("course_id") == str(course["_id"]) and (ex.get("room_id") or "") == str(room["_id"])
                if same:
                    skipped_existing.append(f"{loc} مطابقة تماماً للموجود في النظام — لا تغيير")
                    continue
                replace_id = str(ex["_id"])
                replace_teacher = ex.get("teacher_id", "") or ""
                replace_course = ex.get("course_id", "") or ""
                old_name = courses_by_id.get(ex.get("course_id", ""), "مقرر آخر")
                old_room = rooms_by_id.get(ex.get("room_id", ""), "")
                replace_desc = f"{loc} سيُستبدل '{old_name}'{f' ({old_room})' if old_room else ''} ← بـ'{course.get('name', '')}' ({room.get('name', '')})"

            to_create.append({
                "faculty_id": faculty_id,
                "department_id": department_id,
                "level": level,
                "section": section_val,
                "day": day,
                "slot_number": slot_number,
                "course_id": str(course["_id"]),
                "teacher_id": effective_tid,
                "room_id": str(room["_id"]),
                "_loc": loc,
                "_course_name": course.get("name", ""),
                "_teacher_name": teacher.get("full_name", ""),
                "_room_name": room.get("name", ""),
                "_replace_id": replace_id,
                "_replace_desc": replace_desc,
                "_replaced_teacher": replace_teacher,
                "_replaced_course": replace_course,
                "_cross_level": cross_level,
            })
        r += 3

    replaced_msgs = [it["_replace_desc"] for it in to_create if it["_replace_id"]]
    replaced_ids = {it["_replace_id"] for it in to_create if it["_replace_id"]}

    # 🧭 الملف هو الأساس حرفياً: خلايا المقررات المذكورة في الملف بمواضع غير مذكورة فيه → تُزال (إعادة تموضع)
    removal_ids = set()
    removal_slots = {}
    reposition_msgs = []
    for s in existing_slots:
        cid0 = s.get("course_id", "")
        if cid0 not in file_positions:
            continue
        sid0 = str(s["_id"])
        if sid0 in replaced_ids:
            continue
        if (s.get("day"), s.get("slot_number")) not in file_positions[cid0]:
            removal_ids.add(sid0)
            removal_slots[sid0] = s
            grp = f"م{s.get('level')}{' شعبة ' + s.get('section') if s.get('section') else ''}"
            reposition_msgs.append(
                f"↪️ إعادة تموضع '{courses_by_id.get(cid0, '')}' ({grp}): ستُزال خليته في {s.get('day')} فترة {s.get('slot_number')} — غير مذكورة في الملف (الملف هو الأساس)"
            )

    # تناقض الإسناد: نفس المقرر بأكثر من اسم أستاذ داخل الملف
    for cid, names in file_teachers.items():
        if len(names) > 1:
            conflicts.append(f"تناقض داخل الملف: المقرر '{courses_by_id.get(cid, '')}' مذكور بأكثر من أستاذ ({'، '.join(names.values())}) — وحّد الاسم ثم أعد الرفع")

    # ما بعد المرور: طبّق الإسنادات الجديدة على كل خلايا الملف لنفس المقرر (حتى المذكورة قبل اكتشاف الإسناد)
    for it in to_create:
        if it["course_id"] in reassign_map:
            it["teacher_id"] = reassign_map[it["course_id"]]["new_id"]
            it["_teacher_name"] = reassign_map[it["course_id"]]["new_name"]

    # ===== 🔗 كشف المحاضرات المشتركة داخل الملف: نفس (اليوم/الفترة/المدرس/القاعة/الاسم الأساسي للمقرر) لمستويات/شعب مختلفة =====
    def _base_cname(name, sec):
        n = (name or "").strip()
        suffix = f"({sec})" if sec else ""
        if suffix and n.endswith(suffix):
            n = n[: -len(suffix)].strip()
        return _norm(n)

    merge_msgs = []
    _file_groups = {}
    for it in to_create:
        k = (it["day"], it["slot_number"], it["teacher_id"], it["room_id"], _base_cname(it["_course_name"], it.get("section") or ""))
        _file_groups.setdefault(k, []).append(it)
    for _k, _items in _file_groups.items():
        if len(_items) < 2:
            continue
        _gid = uuid.uuid4().hex
        for it in _items:
            it["merge_group_id"] = _gid
            it["merge_key"] = f"{_gid}:{it['department_id']}:{it['level']}:{it.get('section', '') or ''}"
        _labels = " + ".join(f"م{it['level']}" + (f"/{it['section']}" if it.get("section") else "") for it in _items)
        merge_msgs.append(f"🔗 محاضرة مشتركة: '{_items[0]['_course_name']}' — {_labels} ({_items[0]['day']} الفترة {_items[0]['slot_number']}) بمدرس وقاعة موحدين")

    # ===== 4) فحص التعارضات (داخل الملف + مع الجدول القائم عبر كل الأقسام) =====
    # الخلايا المستبدلة تُستثنى، والإسنادات الجديدة تسري على المحاضرات القائمة لنفس المقرر (تحديث متسلسل)
    all_slots = await db.weekly_schedule.find({}, {"teacher_id": 1, "room_id": 1, "day": 1, "slot_number": 1, "department_id": 1, "level": 1, "section": 1, "course_id": 1, "merge_group_id": 1}).to_list(20000)
    all_slots = [s for s in all_slots if str(s["_id"]) not in replaced_ids and str(s["_id"]) not in removal_ids]
    reassign_new = {cid: info["new_id"] for cid, info in reassign_map.items()}

    pref_tids = list({x["teacher_id"] for x in to_create} | set(reassign_new.values()))
    prefs_map = {p["teacher_id"]: p for p in await db.teacher_preferences.find({"teacher_id": {"$in": pref_tids}}).to_list(500)} if pref_tids else {}

    # 🔧 اعتماد الإسناد الحالي للمقرر (لا معرف المعلم القديم المخزن في الخلية) لمنع تعارضات وهمية من بيانات قديمة
    all_courses_docs = await db.courses.find({}, {"teacher_id": 1, "name": 1, "level": 1, "section": 1, "department_id": 1}).to_list(50000)
    course_info = {str(c["_id"]): c for c in all_courses_docs}
    dept_names = {str(d["_id"]): d.get("name", "") for d in await db.departments.find({}, {"name": 1}).to_list(2000)}

    def _eff_tid(s):
        cid = s.get("course_id", "")
        if cid in reassign_new:
            return reassign_new[cid]
        ct = (course_info.get(cid) or {}).get("teacher_id") or ""
        return ct or s.get("teacher_id")

    def _slot_desc(s):
        c = course_info.get(s.get("course_id", ""), {})
        grp = f"م{c.get('level') or s.get('level') or '?'}" + (f"/{c.get('section')}" if c.get("section") else "")
        dep = dept_names.get(s.get("department_id", ""), "")
        return f"'{c.get('name', 'مقرر آخر')}' ({grp}{' — قسم ' + dep if dep else ''})"

    busy_teacher_owner = {}
    busy_room_owner = {}
    teacher_daily = {}
    _counted_g = set()  # 🔗 أعضاء المحاضرة المشتركة يُحسبون محاضرة واحدة في العبء اليومي
    for s in all_slots:
        if s.get("room_id"):
            busy_room_owner.setdefault((s.get("room_id"), s.get("day"), s.get("slot_number")), s)
        tid = _eff_tid(s)
        if not tid:
            continue
        k = (tid, s.get("day"), s.get("slot_number"))
        if k in busy_teacher_owner:
            other = busy_teacher_owner[k]
            # تصادم قائم بسبب إسناد جديد (أياً كان ترتيب المرور)
            if s.get("course_id", "") in reassign_new or other.get("course_id", "") in reassign_new:
                tname = teachers_map.get(tid, {}).get("full_name", "")
                conflicts.append(f"الإسناد الجديد للأستاذ '{tname}' يجعل محاضرتين قائمتين تتصادمان يوم {s.get('day')} الفترة {s.get('slot_number')}: {_slot_desc(s)} و{_slot_desc(other)}")
        else:
            busy_teacher_owner[k] = s
        dk = (tid, s.get("day"))
        mg0 = s.get("merge_group_id")
        if mg0:
            if (dk, mg0) in _counted_g:
                continue
            _counted_g.add((dk, mg0))
        teacher_daily[dk] = teacher_daily.get(dk, 0) + 1

    # تحقق أن المحاضرات القائمة للمقررات المعاد إسنادها تحترم تفضيلات الأستاذ الجديد وحدّه اليومي
    for s in all_slots:
        cid = s.get("course_id", "")
        if cid not in reassign_new:
            continue
        tid = reassign_new[cid]
        pref = prefs_map.get(tid)
        tname = teachers_map.get(tid, {}).get("full_name", "")
        if pref and _is_period_unavailable(pref, s.get("day"), s.get("slot_number")):
            conflicts.append(f"الإسناد الجديد: '{tname}' غير متاح يوم {s.get('day')} الفترة {s.get('slot_number')} لمحاضرة قائمة لمقرر '{courses_by_id.get(cid, '')}'")
    for tid in set(reassign_new.values()):
        pref = prefs_map.get(tid)
        if not pref:
            continue
        max_daily = int(pref.get("max_daily_lectures") or 3)
        tname = teachers_map.get(tid, {}).get("full_name", "")
        for (t, d), cnt in teacher_daily.items():
            if t == tid and cnt > max_daily:
                conflicts.append(f"الإسناد الجديد: '{tname}' سيتجاوز الحد اليومي ({max_daily}) يوم {d} بالمحاضرات القائمة")

    seen_teacher, seen_room, seen_cell = {}, {}, set()
    counted_groups = set()
    existing_merge_joins = {}  # 🔗 محاضرات قائمة ستتحول لمشتركة بانضمام خلايا الملف إليها

    def _is_same_lecture(exslot, item):
        exc = course_info.get(exslot.get("course_id", ""), {})
        return (
            (exslot.get("room_id") or "") == item["room_id"]
            and _eff_tid(exslot) == item["teacher_id"]
            and _base_cname(exc.get("name", ""), exc.get("section") or "") == _base_cname(item["_course_name"], item.get("section") or "")
        )

    for item in to_create:
        loc = item["_loc"]
        tk = (item["teacher_id"], item["day"], item["slot_number"])
        rk = (item["room_id"], item["day"], item["slot_number"])
        ck = (item["level"], item["section"], item["day"], item["slot_number"])
        if ck in seen_cell:
            conflicts.append(f"{loc} خلية مكررة داخل الملف لنفس الشعبة")
        gid_i = item.get("merge_group_id", "")
        # 🔗 خلية بمقرر من مستوى آخر: تُقبل فقط إذا كانت جزءاً من محاضرة مشتركة (ملف أو نظام)
        if item.get("_cross_level") and not gid_i:
            bt0 = busy_teacher_owner.get(tk)
            if not (bt0 and _is_same_lecture(bt0, item)):
                errors.append(
                    f"{loc} المقرر '{item['_course_name']}' يخص مستوى آخر — يُقبل فقط كمحاضرة مشتركة: "
                    f"ضع نفس المحاضرة (نفس المدرس والقاعة) في نفس اليوم/الفترة لمستواه الأصلي أيضاً — تُخُطيت الخلية"
                )
                item["_dropped"] = True
                continue
        joined_existing = None
        bt = busy_teacher_owner.get(tk)
        if bt:
            if _is_same_lecture(bt, item):
                joined_existing = bt  # 🔗 نفس المحاضرة قائمة لمستوى/شعبة أخرى → دمج بدل التعارض
            else:
                conflicts.append(f"{loc} تعارض معلم: '{item['_teacher_name']}' مشغول بمحاضرة قائمة {_slot_desc(bt)} بنفس (اليوم/الفترة)")
        elif tk in seen_teacher and (not gid_i or seen_teacher[tk] != gid_i):
            conflicts.append(f"{loc} تعارض معلم داخل الملف: '{item['_teacher_name']}' مذكور في خليتين بنفس (اليوم/الفترة)")
        br = busy_room_owner.get(rk)
        if br:
            if not _is_same_lecture(br, item):
                conflicts.append(f"{loc} تعارض قاعة: '{item['_room_name']}' محجوزة لمحاضرة قائمة {_slot_desc(br)} بنفس (اليوم/الفترة)")
        elif rk in seen_room and (not gid_i or seen_room[rk] != gid_i):
            conflicts.append(f"{loc} تعارض قاعة داخل الملف: '{item['_room_name']}' مذكورة في خليتين بنفس (اليوم/الفترة)")
        if joined_existing is not None:
            gid = joined_existing.get("merge_group_id") or uuid.uuid4().hex
            members = [x for x in to_create if gid_i and x.get("merge_group_id") == gid_i] or [item]
            for x in members:
                x["merge_group_id"] = gid
                x["merge_key"] = f"{gid}:{x['department_id']}:{x['level']}:{x.get('section', '') or ''}"
            gid_i = gid
            if not joined_existing.get("merge_group_id"):
                joined_existing["merge_group_id"] = gid
                existing_merge_joins[str(joined_existing["_id"])] = (joined_existing, gid)
            merge_msgs.append(f"🔗 {loc} سينضم لمحاضرة قائمة {_slot_desc(joined_existing)} كمحاضرة مشتركة")
        pref = prefs_map.get(item["teacher_id"])
        # استبدال محايد زمنياً: نفس المعلم الفعلي كان يشغل الخلية أصلاً — لا يُعامل كمحاضرة إضافية
        rep_eff = ""
        if item["_replace_id"]:
            rc = item.get("_replaced_course", "")
            rep_eff = reassign_new.get(rc) or (course_info.get(rc, {}) or {}).get("teacher_id") or item.get("_replaced_teacher", "")
        neutral_time = bool(item["_replace_id"]) and rep_eff == item["teacher_id"]
        if pref and not neutral_time and joined_existing is None and _is_period_unavailable(pref, item["day"], item["slot_number"]):
            conflicts.append(f"{loc} تعارض تفضيلات: '{item['_teacher_name']}' غير متاح يوم {item['day']} الفترة {item['slot_number']}")
        dk = (item["teacher_id"], item["day"])
        already_counted = (joined_existing is not None) or (gid_i and (dk, gid_i) in counted_groups)
        if not already_counted:
            teacher_daily[dk] = teacher_daily.get(dk, 0) + 1
            if gid_i:
                counted_groups.add((dk, gid_i))
            if pref and not neutral_time and teacher_daily[dk] > int(pref.get("max_daily_lectures") or 3):
                conflicts.append(f"{loc} تعارض تفضيلات: '{item['_teacher_name']}' سيتجاوز الحد اليومي ({pref.get('max_daily_lectures', 3)}) يوم {item['day']}")
        seen_teacher[tk] = gid_i
        seen_room[rk] = gid_i
        seen_cell.add(ck)

    # إسقاط الخلايا المرفوضة (مستوى آخر بدون دمج) وإعادة حساب الاستبدالات
    if any(x.get("_dropped") for x in to_create):
        to_create = [x for x in to_create if not x.get("_dropped")]
        replaced_msgs = [it["_replace_desc"] for it in to_create if it["_replace_id"]]
        replaced_ids = {it["_replace_id"] for it in to_create if it["_replace_id"]}

    new_count = sum(1 for it in to_create if not it["_replace_id"])
    reassign_msgs = [
        (f"المقرر '{info['course_name']}': كان بلا إسناد ← سيُسند إلى '{info['new_name']}'" if not info["old_id"]
         else f"المقرر '{info['course_name']}': الإسناد سيتغير من '{info['old_name']}' ← إلى '{info['new_name']}' (يسري على كل محاضراته)")
        for info in reassign_map.values()
    ]
    can_commit = len(conflicts) == 0 and (len(to_create) > 0 or len(reassign_map) > 0 or len(removal_ids) > 0)
    report = {
        "dry_run": is_dry,
        "to_create": new_count,
        "to_replace": len(replaced_msgs),
        "replaced": replaced_msgs,
        "to_reassign": len(reassign_msgs),
        "reassigned": reassign_msgs,
        "to_reposition": len(removal_ids),
        "repositioned": reposition_msgs,
        "to_merge": len(merge_msgs),
        "merged": merge_msgs,
        "created": 0,
        "skipped_existing": skipped_existing,
        "errors": errors,
        "conflicts": conflicts,
        "can_commit": can_commit,
    }

    if conflicts:
        report["message"] = f"🛑 تم إيقاف الاستيراد: يوجد {len(conflicts)} تعارض جدولة — عالج التعارضات ثم أعد المحاولة (لم يُحفظ أي شيء)"
        return report
    if is_dry:
        if not to_create and not reassign_map and not removal_ids:
            report["message"] = (
                "الملف مطابق تماماً للجدول الحالي — لا توجد تغييرات للاستيراد"
                + (f" ({len(skipped_existing)} خلية مطابقة)" if skipped_existing else "")
                + (f" • {len(errors)} خطأ أسماء (انظر التقرير)" if errors else "")
            )
            return report
        report["message"] = (
            f"معاينة: سيتم إدراج {new_count} محاضرة جديدة"
            + (f" • 🔗 {len(merge_msgs)} دمج كمحاضرات مشتركة" if merge_msgs else "")
            + (f" • استبدال {len(replaced_msgs)} خلية بمحتوى الملف" if replaced_msgs else "")
            + (f" • إعادة تموضع: إزالة {len(removal_ids)} خلية غير مذكورة في الملف" if removal_ids else "")
            + (f" • تغيير إسناد {len(reassign_msgs)} مقرر" if reassign_msgs else "")
            + (f" • {len(skipped_existing)} مطابقة بلا تغيير" if skipped_existing else "")
            + (f" • {len(errors)} خطأ أسماء" if errors else "")
        )
        return report

    # ===== 5) تنفيذ: حذف المستبدلة والمعاد تموضعها ← تطبيق الإسنادات ← إدراج محاضرات الملف =====
    # 🔄 مزامنة المحاضرات اليومية المستقبلية للخلايا التي ستُزال (استبدال أو إعادة تموضع)
    lec_sync = {"deleted": 0, "skipped_attendance": 0, "skipped_rescheduled": 0}
    replaced_slots = {str(s["_id"]): s for s in existing_slots if str(s["_id"]) in replaced_ids}
    for sdoc in list(removal_slots.values()) + list(replaced_slots.values()):
        try:
            res = await _sync_future_lectures(db, sdoc, "delete")
            for k in lec_sync:
                lec_sync[k] += res.get(k, 0)
        except Exception:
            pass

    all_remove_ids = replaced_ids | removal_ids
    if all_remove_ids:
        await db.weekly_schedule.delete_many({"_id": {"$in": [ObjectId(x) for x in all_remove_ids]}})

    for cid, info in reassign_map.items():
        try:
            await db.weekly_schedule.update_many({"course_id": cid}, {"$set": {"teacher_id": info["new_id"]}})
        except DuplicateKeyError:
            raise HTTPException(status_code=409, detail=f"تعذر تغيير إسناد '{info['course_name']}': تصادم في جدول الأستاذ '{info['new_name']}' — أعد المعاينة")
        await db.courses.update_one({"_id": ObjectId(cid)}, {"$set": {"teacher_id": info["new_id"]}})
        # ⚖️ العبء التدريسي يتبع الإسناد الجديد
        if info.get("old_id"):
            await db.teaching_loads.delete_many({"course_id": cid, "teacher_id": info["old_id"]})
        if info["new_id"] and not await db.teaching_loads.find_one({"course_id": cid, "teacher_id": info["new_id"]}):
            course_doc = await db.courses.find_one({"_id": ObjectId(cid)})
            load = {
                "teacher_id": info["new_id"], "course_id": cid,
                "weekly_hours": (course_doc or {}).get("credit_hours", 3),
                "created_by": current_user.get("id", ""),
                "created_at": datetime.now(timezone.utc), "updated_at": datetime.now(timezone.utc),
            }
            active_sem = await db.semesters.find_one({"status": "active"})
            sem_id = str(active_sem["_id"]) if active_sem else (course_doc or {}).get("semester_id")
            if sem_id:
                load["semester_id"] = sem_id
            await db.teaching_loads.insert_one(load)

    created = 0
    replaced_count = 0
    for item in to_create:
        doc = {k: v for k, v in item.items() if not k.startswith("_")}
        doc["created_at"] = datetime.now(timezone.utc)
        doc["created_by"] = current_user["id"]
        doc["imported_from_excel"] = True
        if item["_replace_id"]:
            doc["replaced_existing"] = True
        try:
            await db.weekly_schedule.insert_one(doc)
            if item["_replace_id"]:
                replaced_count += 1
            else:
                created += 1
        except DuplicateKeyError:
            conflicts.append(f"{item['_loc']} رُفض من قاعدة البيانات (تعارض فريد لحظي)")

    # 🔗 تحويل المحاضرات القائمة التي انضمت إليها خلايا الملف إلى مشتركة
    for _sid, (_sdoc, _gid) in existing_merge_joins.items():
        await db.weekly_schedule.update_one({"_id": ObjectId(_sid)}, {"$set": {
            "merge_group_id": _gid, "merge_key": f"{_gid}:{_sdoc.get('department_id', '')}:{_sdoc.get('level')}:{_sdoc.get('section', '') or ''}"}})

    await log_activity(
        current_user, "import_master_schedule_excel", "weekly_schedule", department_id, None,
        {"faculty_id": faculty_id, "created": created, "replaced": replaced_count, "repositioned": len(removal_ids), "reassigned": len(reassign_map), "errors": len(errors), "skipped_identical": len(skipped_existing)},
    )
    report["created"] = created
    report["replaced_count"] = replaced_count
    report["message"] = (
        f"✅ تم إدراج {created} محاضرة جديدة"
        + (f" • 🔗 {len(merge_msgs)} دمج كمحاضرات مشتركة" if merge_msgs else "")
        + (f" • استُبدلت {replaced_count} خلية بمحتوى الملف" if replaced_count else "")
        + (f" • أُزيلت {len(removal_ids)} خلية غير مذكورة في الملف (إعادة تموضع)" if removal_ids else "")
        + (f" • تغيّر إسناد {len(reassign_map)} مقرر" if reassign_map else "")
        + (f" • {len(skipped_existing)} مطابقة بلا تغيير" if skipped_existing else "")
        + (f" • {len(errors)} خلية بأخطاء أسماء (انظر التقرير)" if errors else "")
        + (f" • 📅 حُذفت {lec_sync['deleted']} محاضرة يومية مستقبلية للخلايا المزالة" if lec_sync["deleted"] else "")
        + (f" (استُثنيت {lec_sync['skipped_attendance'] + lec_sync['skipped_rescheduled']}: حضور/معاد جدولتها)" if (lec_sync["skipped_attendance"] + lec_sync["skipped_rescheduled"]) else "")
    )
    return report
