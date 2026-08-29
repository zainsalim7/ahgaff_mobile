"""نظام الدرجات: استيراد كشوفات الإكسل الفصلية + السجل الأكاديمي التراكمي + بيان حالة ودرجات PDF مع QR تحقق"""
import io
import re
import uuid
from datetime import datetime, timezone
from typing import Optional, List

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from bson import ObjectId

from .deps import get_db, get_current_user, log_activity, has_permission

router = APIRouter()

TERM_AR = {1: "الأول", 2: "الثاني", 3: "الصيفي"}
LEVEL_AR = {1: "الأول", 2: "الثاني", 3: "الثالث", 4: "الرابع", 5: "الخامس", 6: "السادس"}


def _can_manage(user: dict) -> bool:
    if user.get("role") in ("admin", "registrar", "registration_manager"):
        return True
    return has_permission(user, "manage_grades")


def _norm_name(n: str) -> str:
    n = str(n or "").strip()
    n = re.sub(r"[ـ]", "", n)
    n = re.sub(r"[أإآٱ]", "ا", n)
    n = n.replace("ة", "ه").replace("ى", "ي").replace("ئ", "ي").replace("ؤ", "و")
    n = re.sub(r"\s+", " ", n)
    return n


def _fmt_val(v):
    if v is None:
        return ""
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else str(round(v, 2))
    return str(v).strip()


class _Sheet:
    """موحّد قراءة لأوراق xls (xlrd) و xlsx (openpyxl)"""

    def __init__(self, nrows, ncols, getter):
        self.nrows, self.ncols, self._get = nrows, ncols, getter

    def cell(self, r, c):
        if r >= self.nrows or c >= self.ncols:
            return ""
        return self._get(r, c)


def _open_workbook(content: bytes, filename: str) -> dict:
    """يرجع {sheet_name: _Sheet}"""
    sheets = {}
    if filename.lower().endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(file_contents=content)
        for name in wb.sheet_names():
            sh = wb.sheet_by_name(name)
            sheets[name] = _Sheet(sh.nrows, sh.ncols, lambda r, c, _s=sh: _s.cell_value(r, c))
    else:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            nrows = len(rows)
            ncols = max((len(r) for r in rows), default=0)
            sheets[name] = _Sheet(nrows, ncols, lambda r, c, _rows=rows: (_rows[r][c] if c < len(_rows[r]) else "") or "")
    return sheets


# ===== 📄 النموذج الموحّد =====

TEMPLATE_META_KEYS = {"الكلية": "faculty_name", "القسم": "department_name", "التخصص": "department_name",
                      "المستوى": "level", "الفصل": "semester_no", "العام الجامعي": "academic_year", "الدفعة": "batch_no"}


@router.get("/grades/template")
async def download_grades_template(current_user: dict = Depends(get_current_user)):
    """⬇️ تنزيل نموذج كشف الدرجات الموحّد (Excel)"""
    if not _can_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "كشف الدرجات"
    ws.sheet_view.rightToLeft = True
    bold = Font(bold=True, size=11)
    hdr_fill = PatternFill("solid", fgColor="E8F0E8")
    meta_fill = PatternFill("solid", fgColor="FFF6E0")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(*[Side(style="thin")] * 4)

    ws["A1"] = "نموذج كشف درجات فصلي — جامعة الأحقاف (لا تغيّر أسماء الحقول)"
    ws["A1"].font = Font(bold=True, size=12, color="1a5c2a")
    meta_rows = [("الكلية", "كلية الشريعة والقانون"), ("القسم", "شريعة"), ("المستوى", 4),
                 ("الفصل", 1), ("العام الجامعي", "2025-2026"), ("الدفعة", 27)]
    for i, (k, v) in enumerate(meta_rows, start=2):
        ws.cell(row=i, column=1, value=k).font = bold
        ws.cell(row=i, column=1).fill = meta_fill
        ws.cell(row=i, column=1).border = thin
        ws.cell(row=i, column=2, value=v).border = thin

    hr = 9  # صف رؤوس الجدول
    ws.cell(row=hr, column=1, value="رقم القيد")
    ws.cell(row=hr, column=2, value="اسم الطالب")
    sample_courses = ["أصول الفقه (4)", "البلاغة (3)", "مدخل القانون (3)"]
    col = 3
    for cn in sample_courses:
        ws.cell(row=hr, column=col, value=cn)
        ws.cell(row=hr + 1, column=col, value="الدرجة")
        ws.cell(row=hr + 1, column=col + 1, value="التقدير")
        ws.merge_cells(start_row=hr, start_column=col, end_row=hr, end_column=col + 1)
        col += 2
    ws.cell(row=hr, column=col, value="النتيجة")
    ws.cell(row=hr, column=col + 1, value="ملاحظات")
    for c in range(1, col + 2):
        for r in (hr, hr + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = bold
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = thin
    # صفا مثال
    ws.cell(row=hr + 2, column=1, value="ش129134")
    ws.cell(row=hr + 2, column=2, value="مثال: فلان بن فلان (احذف هذا الصف)")
    ws.cell(row=hr + 2, column=3, value=87)
    ws.cell(row=hr + 2, column=4, value="جيد جداً")
    ws.cell(row=hr + 2, column=5, value=62)
    ws.cell(row=hr + 2, column=7, value=90)
    ws.cell(row=hr + 2, column=col, value=4.15)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=grades_template.xlsx"})


def _parse_template_sheet(sh) -> Optional[dict]:
    """قراءة النموذج الموحّد — يرجع None إن لم يكن الملف بنمط النموذج.
    التمييز الصارم: «رقم القيد» يليه مباشرة «اسم الطالب» (الملفات القديمة تستخدم «الاسم»)"""
    header_r, reg_c = None, None
    for r in range(min(20, sh.nrows)):
        for c in range(min(6, sh.ncols)):
            if str(sh.cell(r, c)).strip() == "رقم القيد" and str(sh.cell(r, c + 1)).strip() == "اسم الطالب":
                header_r, reg_c = r, c
                break
        if header_r is not None:
            break
    if header_r is None:
        return None
    name_c = reg_c + 1
    # البيانات الوصفية أعلى الجدول
    meta = {}
    for r in range(header_r):
        for c in range(min(4, sh.ncols)):
            k = str(sh.cell(r, c)).strip()
            if k in TEMPLATE_META_KEYS:
                meta[TEMPLATE_META_KEYS[k]] = _fmt_val(sh.cell(r, c + 1))
    # أعمدة المقررات (مع أزواج الدرجة/التقدير)
    sub_r = header_r + 1
    has_sub = any(str(sh.cell(sub_r, c)).strip() in ("الدرجة", "التقدير") for c in range(name_c + 1, sh.ncols))
    courses, cols = [], []  # cols: (name, credits, total_col, letter_col|None)
    c = name_c + 1
    while c < sh.ncols:
        v = str(sh.cell(header_r, c)).strip()
        if v in ("النتيجة", "ملاحظات"):
            break
        if v:
            m = re.match(r"^(.*?)\s*\((\d+(?:\.\d+)?)\)\s*$", v)
            nm, cr = (m.group(1).strip(), m.group(2)) if m else (v, "")
            letter_col = None
            if has_sub and c + 1 < sh.ncols and str(sh.cell(sub_r, c + 1)).strip() == "التقدير":
                letter_col = c + 1
            courses.append({"name": nm, "credits": cr})
            cols.append((nm, cr, c, letter_col))
            c += 2 if letter_col is not None else 1
        else:
            c += 1
    result_c = note_c = None
    for cc in range(name_c + 1, sh.ncols):
        v = str(sh.cell(header_r, cc)).strip()
        if v == "النتيجة":
            result_c = cc
        elif v == "ملاحظات":
            note_c = cc
    students, skipped = [], 0
    start = sub_r + 1 if has_sub else header_r + 1
    for r in range(start, sh.nrows):
        nm = str(sh.cell(r, name_c)).strip()
        reg = _fmt_val(sh.cell(r, reg_c))
        if not nm and not reg:
            continue
        if not reg or "احذف هذا الصف" in nm:
            skipped += 1
            continue
        grades = []
        for (cn, cr, tc, lc) in cols:
            g = {"course_name": cn, "credits": cr, "total": _fmt_val(sh.cell(r, tc))}
            if lc is not None:
                letter = _fmt_val(sh.cell(r, lc))
                if letter:
                    g["grade_letter"] = letter
            grades.append(g)
        students.append({"name": nm, "reg_no": reg,
                         "grades": grades,
                         "result": _fmt_val(sh.cell(r, result_c)) if result_c is not None else "",
                         "note": _fmt_val(sh.cell(r, note_c)) if note_c is not None else ""})
    return {"courses": courses, "students": students, "meta": meta, "skipped_no_reg": skipped, "is_template": True}


def _parse_grades_file(content: bytes, filename: str) -> dict:
    """يقرأ ورقة «المساقات» (الاسم + الوحدات) و«السعي والأدوار» (مجموع كل مقرر لكل طالب)
    و«ملخص الدور الأول» (النتيجة/الملاحظات) — بأسلوب مرن"""
    sheets = _open_workbook(content, filename)

    # 🆕 محاولة قراءة النموذج الموحّد أولاً (أي ورقة تحتوي «رقم القيد»)
    for sh0 in sheets.values():
        tpl = _parse_template_sheet(sh0)
        if tpl is not None:
            return tpl

    # 1) المساقات
    courses = []
    saq = next((sheets[n] for n in sheets if "المساقات" in n), None)
    if saq:
        header_r = None
        for r in range(min(15, saq.nrows)):
            for c in range(min(5, saq.ncols)):
                if str(saq.cell(r, c)).strip() == "الرمز":
                    header_r = r
                    break
            if header_r is not None:
                break
        if header_r is not None:
            name_c, units_c = None, None
            for c in range(saq.ncols):
                v = str(saq.cell(header_r, c)).strip()
                if v == "المساق":
                    name_c = c
                elif v == "الوحدات":
                    units_c = c
            if name_c is not None:
                for r in range(header_r + 1, saq.nrows):
                    nm = str(saq.cell(r, name_c)).strip()
                    if not nm:
                        break
                    cr = _fmt_val(saq.cell(r, units_c)) if units_c is not None else ""
                    eff = _fmt_val(saq.cell(r, units_c + 1)) if units_c is not None else ""
                    courses.append({"name": nm, "credits": cr, "effort_max": eff})

    # 2) السعي والأدوار — المجاميع
    sa = next((sheets[n] for n in sheets if "السعي والأدوار" in n), None)
    if not sa:
        raise HTTPException(status_code=400, detail="لم يتم العثور على ورقة «السعي والأدوار» في الملف")
    header_r, name_c = None, None
    for r in range(min(20, sa.nrows)):
        for c in range(min(12, sa.ncols)):
            if str(sa.cell(r, c)).strip() == "الاسم":
                header_r, name_c = r, c
                break
        if header_r is not None:
            break
    if header_r is None:
        raise HTTPException(status_code=400, detail="تعذّر تحديد صف العناوين (عمود «الاسم») في ورقة السعي والأدوار")
    reg_c = name_c - 1  # رقم القيد

    # أعمدة «مجموع» ← اسم المقرر من الصف الأعلى (خلية مدموجة تبدأ قبل 4 أعمدة)
    course_cols = []  # [(course_name, total_col)]
    for c in range(name_c + 1, sa.ncols):
        if str(sa.cell(header_r, c)).strip() == "مجموع":
            nm = ""
            for cc in range(c, max(c - 6, name_c), -1):
                v = str(sa.cell(header_r - 1, cc)).strip()
                if v:
                    nm = v
                    break
            if nm:
                course_cols.append((nm, c))

    credits_map = {ci["name"].strip(): ci["credits"] for ci in courses}

    students = []
    r = header_r + 1
    # تخطي صف «أول/ثان/تصفية»
    while r < sa.nrows and not str(sa.cell(r, name_c)).strip():
        r += 1
    while r < sa.nrows:
        nm = str(sa.cell(r, name_c)).strip()
        if not nm:
            break
        reg = _fmt_val(sa.cell(r, reg_c)) if reg_c >= 0 else ""
        grades = []
        for (cn, tc) in course_cols:
            grades.append({"course_name": cn, "credits": credits_map.get(cn, ""), "total": _fmt_val(sa.cell(r, tc)),
                           "s": _fmt_val(sa.cell(r, tc - 4)), "r1": _fmt_val(sa.cell(r, tc - 3)),
                           "r2": _fmt_val(sa.cell(r, tc - 2)), "tas": _fmt_val(sa.cell(r, tc - 1))})
        students.append({"name": nm, "reg_no": reg, "grades": grades, "result": "", "note": ""})
        r += 1

    # 🆕 استكمال أرقام القيد الفارغة من ورقة «السعي» (المصدر الذي تُنقل منه الأسماء)
    if any(not s["reg_no"] for s in students):
        s3 = next((sheets[n] for n in sheets if n.strip() == "السعي"), None)
        if s3:
            hr3, nc3 = None, None
            for r0 in range(min(20, s3.nrows)):
                for c0 in range(min(12, s3.ncols)):
                    if str(s3.cell(r0, c0)).strip() == "الاسم":
                        hr3, nc3 = r0, c0
                        break
                if hr3 is not None:
                    break
            if hr3 is not None:
                reg_cols = [c0 for c0 in range(nc3) if str(s3.cell(hr3, c0)).strip() == "رقم القيد"]
                reg_map = {}
                for r0 in range(hr3 + 1, s3.nrows):
                    nm0 = str(s3.cell(r0, nc3)).strip()
                    if not nm0:
                        continue
                    reg0 = ""
                    for c0 in reg_cols:
                        v = _fmt_val(s3.cell(r0, c0))
                        if v:
                            reg0 = v
                            break
                    if reg0:
                        reg_map[_norm_name(nm0)] = reg0
                for st in students:
                    if not st["reg_no"]:
                        st["reg_no"] = reg_map.get(_norm_name(st["name"]), "")

    # 3) ملخص الدور الأول — النتيجة والملاحظات (اختياري، مطابقة بالاسم)
    summ = next((sheets[n] for n in sheets if n.strip() == "ملخص الدور الأول"), None)
    if summ and summ.ncols >= 3:
        s_header, s_name_c = None, None
        for r0 in range(min(15, summ.nrows)):
            for c0 in range(min(6, summ.ncols)):
                if str(summ.cell(r0, c0)).strip() == "الاسم":
                    s_header, s_name_c = r0, c0
                    break
            if s_header is not None:
                break
        if s_header is not None:
            res_c, note_c = summ.ncols - 2, summ.ncols - 1
            smap = {}
            for r0 in range(s_header + 1, summ.nrows):
                nm0 = str(summ.cell(r0, s_name_c)).strip()
                if not nm0 or "إجمالي" in nm0 or "نسبة" in nm0:
                    continue
                res = _fmt_val(summ.cell(r0, res_c))
                try:
                    fres = float(res)
                    res = str(round(fres, 2))
                except (ValueError, TypeError):
                    pass
                smap[_norm_name(nm0)] = (res, _fmt_val(summ.cell(r0, note_c)))
            for st in students:
                hit = smap.get(_norm_name(st["name"]))
                if hit:
                    st["result"], st["note"] = hit

    effort_map = {ci["name"].strip(): ci.get("effort_max", "") for ci in courses}
    return {"courses": [{"name": c[0], "credits": credits_map.get(c[0], ""), "effort_max": effort_map.get(c[0], "")} for c in course_cols], "students": students}


# ===== 📊 تحليل النتيجة (متدرج حسب مرحلة الرصد) =====

def _num(v):
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return None


async def _compute_analysis(db, import_id: str) -> dict:
    """تحليل متدرج لكشف مستورد: سعي فقط ← سعي+دور أول ← نتيجة مكتملة"""
    imp = await db.grade_imports.find_one({"_id": ObjectId(import_id)})
    if not imp:
        raise HTTPException(status_code=404, detail="الاستيراد غير موجود")
    recs = await db.student_grades.find({"import_id": import_id}).to_list(2000)
    courses = imp.get("courses", [])
    has_details = any(("s" in g or "r1" in g) for r in recs for g in r.get("grades", []))

    # كشف المرحلة
    r1_filled = sum(1 for r in recs for g in r.get("grades", []) if _num(g.get("r1")) is not None)
    r2_filled = sum(1 for r in recs for g in r.get("grades", []) if _num(g.get("r2")) is not None or _num(g.get("tas")) is not None)
    stage = "saai" if r1_filled == 0 else ("final" if r2_filled > 0 else "round1")
    if not has_details:
        stage = "final"

    PASS = 50.0
    course_stats, fail_lists = [], {}
    for c in courses:
        cn = c["name"]
        eff_max = _num(c.get("effort_max")) or 0
        vals, saais, fails, absents = [], [], [], 0
        for r in recs:
            g = next((x for x in r.get("grades", []) if x.get("course_name") == cn), None)
            if not g:
                continue
            s = _num(g.get("s"))
            if s is not None:
                saais.append(s)
            tot = _num(g.get("total"))
            raw = str(g.get("total", "")).strip()
            if raw in ("غ", "غش"):
                absents += 1
            if tot is not None:
                vals.append(tot)
                if stage != "saai" and tot < PASS:
                    fails.append(r.get("student_name", ""))
        use = saais if (stage == "saai" and saais) else vals
        st = {"course": cn, "credits": c.get("credits", ""), "count": len(use),
              "avg": round(sum(use) / len(use), 1) if use else 0,
              "max": max(use) if use else 0, "min": min(use) if use else 0,
              "absent": absents}
        if stage == "saai":
            st["effort_max"] = eff_max
            st["weak"] = sum(1 for v in saais if eff_max and v < eff_max / 2)
        else:
            st["fail_count"] = len(fails)
            st["pass_rate"] = round(100 * (len(vals) - len(fails)) / len(vals), 1) if vals else 0
            st["avg_saai"] = round(sum(saais) / len(saais), 1) if saais else None
            fail_lists[cn] = fails
        course_stats.append(st)

    students_out, passed, second_round, dismissed = [], 0, [], []
    saai_alerts = []
    eff_maxes = {c["name"]: _num(c.get("effort_max")) or 0 for c in courses}
    for r in recs:
        gs = r.get("grades", [])
        failed_courses = [g["course_name"] for g in gs if _num(g.get("total")) is not None and _num(g["total"]) < PASS]
        avg_v = _num(r.get("result"))
        item = {"name": r.get("student_name", ""), "reg_no": r.get("reg_no", ""),
                "result": r.get("result", ""), "note": r.get("note", ""),
                "failed_courses": failed_courses, "avg": avg_v}
        students_out.append(item)
        if stage == "saai":
            weak = [g["course_name"] for g in gs if eff_maxes.get(g["course_name"]) and _num(g.get("s")) is not None and _num(g["s"]) < eff_maxes[g["course_name"]] / 2]
            if weak:
                saai_alerts.append({"name": item["name"], "reg_no": item["reg_no"], "courses": weak})
        else:
            if failed_courses:
                second_round.append({"name": item["name"], "reg_no": item["reg_no"], "courses": failed_courses})
            else:
                passed += 1
            if "فصل" in (r.get("note") or "") or "فصل" in (r.get("result") or ""):
                dismissed.append({"name": item["name"], "note": r.get("note", "")})

    top = sorted([s for s in students_out if s["avg"] is not None], key=lambda x: -x["avg"])[:10]
    hardest = min([c for c in course_stats if "pass_rate" in c], key=lambda x: x["pass_rate"], default=None)

    return {
        "stage": stage,
        "has_details": has_details,
        "label": f"{imp.get('department_id', '')}",
        "info": {"filename": imp.get("filename", ""), "level": imp.get("level"),
                 "semester_no": imp.get("semester_no"), "academic_year": imp.get("academic_year", ""),
                 "batch_no": imp.get("batch_no", ""), "students": len(recs)},
        "course_stats": course_stats,
        "summary": {"passed": passed, "second_round": len(second_round), "dismissed": len(dismissed)},
        "top": top,
        "second_round": second_round,
        "fail_lists": fail_lists,
        "saai_alerts": saai_alerts,
        "dismissed": dismissed,
    }


@router.get("/grades/analysis/{import_id}")
async def analyze_import(import_id: str, current_user: dict = Depends(get_current_user)):
    if not _can_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    return await _compute_analysis(get_db(), import_id)


@router.get("/grades/analysis-cumulative")
async def cumulative_analysis(current_user: dict = Depends(get_current_user)):
    """تقارير تراكمية عبر كل الاستيرادات: مقارنة الكشوف + الرسوب المتكرر"""
    if not _can_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    imports = await db.grade_imports.find().to_list(200)
    recs = await db.student_grades.find().to_list(10000)
    comp = []
    for imp in imports:
        iid = str(imp["_id"])
        rs = [r for r in recs if r.get("import_id") == iid]
        avgs = [v for v in (_num(r.get("result")) for r in rs) if v is not None]
        fails = sum(1 for r in rs if any(_num(g.get("total")) is not None and _num(g["total"]) < 50 for g in r.get("grades", [])))
        comp.append({"filename": imp.get("filename", ""), "level": imp.get("level"),
                     "semester_no": imp.get("semester_no"), "academic_year": imp.get("academic_year", ""),
                     "batch_no": imp.get("batch_no", ""), "students": len(rs),
                     "avg": round(sum(avgs) / len(avgs), 2) if avgs else None,
                     "fail_students": fails})
    # الرسوب المتكرر: نفس الطالب رسب في نفس المقرر في أكثر من كشف
    fail_map: dict = {}
    for r in recs:
        key = r.get("reg_no", "")
        if not key:
            continue
        for g in r.get("grades", []):
            t = _num(g.get("total"))
            if t is not None and t < 50:
                fail_map.setdefault((key, r.get("student_name", ""), g["course_name"]), 0)
                fail_map[(key, r.get("student_name", ""), g["course_name"])] += 1
    repeats = [{"reg_no": k[0], "name": k[1], "course": k[2], "times": v}
               for k, v in fail_map.items() if v > 1]
    repeats.sort(key=lambda x: -x["times"])
    return {"comparison": sorted(comp, key=lambda x: (x["academic_year"], x["semester_no"] or 0)),
            "repeat_failures": repeats[:100]}


# ===== الاستيراد =====

@router.post("/grades/import")
async def import_grades(
    file: UploadFile = File(...),
    faculty_id: str = Form(""),
    department_id: str = Form(""),
    level: int = Form(0),
    semester_no: int = Form(0),
    academic_year: str = Form(""),
    batch_no: str = Form(""),
    commit: str = Form("false"),
    replace: str = Form("false"),
    current_user: dict = Depends(get_current_user),
):
    """معاينة (commit=false) أو اعتماد (commit=true) استيراد كشف درجات فصلي من Excel"""
    if not _can_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    content = await file.read()
    parsed = _parse_grades_file(content, file.filename or "grades.xls")
    if not parsed["students"]:
        raise HTTPException(status_code=400, detail="لم يتم العثور على أي طالب في الملف")

    # 🆕 حلّ البيانات الوصفية من النموذج الموحّد (الكلية/القسم بالاسم + المستوى/الفصل/العام/الدفعة)
    meta = parsed.get("meta") or {}
    resolved = {}
    if meta.get("department_name"):
        dq = await db.departments.find_one({"name": {"$regex": re.escape(meta["department_name"].strip()), "$options": "i"}})
        if dq:
            resolved["department_id"] = str(dq["_id"])
            resolved["faculty_id"] = dq.get("faculty_id", "")
            resolved["department_name"] = dq.get("name", "")
    if meta.get("faculty_name") and not resolved.get("faculty_id"):
        fq = await db.faculties.find_one({"name": {"$regex": re.escape(meta["faculty_name"].strip()), "$options": "i"}})
        if fq:
            resolved["faculty_id"] = str(fq["_id"])
    for k in ("level", "semester_no"):
        if meta.get(k):
            try:
                resolved[k] = int(float(meta[k]))
            except (ValueError, TypeError):
                pass
    for k in ("academic_year", "batch_no"):
        if meta.get(k):
            resolved[k] = str(meta[k])
    # قيم الملف تكمل ما لم يُدخل — والملف النموذجي مصدر الحقيقة عند المعاينة
    if parsed.get("is_template"):
        faculty_id = resolved.get("faculty_id") or faculty_id
        department_id = resolved.get("department_id") or department_id
        level = resolved.get("level") or level
        semester_no = resolved.get("semester_no") or semester_no
        academic_year = resolved.get("academic_year") or academic_year
        batch_no = resolved.get("batch_no") or batch_no
    if commit == "true" and not (faculty_id and department_id and level and semester_no and academic_year):
        raise HTTPException(status_code=400, detail="أكمل بيانات الكشف (الكلية/القسم/المستوى/الفصل/العام الجامعي)")

    # مطابقة الطلاب: رقم القيد أولاً ثم الاسم
    all_students = await db.students.find({}, {"student_id": 1, "full_name": 1, "department_id": 1}).to_list(20000)
    by_reg = {str(s.get("student_id", "")).strip(): s for s in all_students if s.get("student_id")}
    by_name: dict = {}
    for s in all_students:
        by_name.setdefault(_norm_name(s.get("full_name", "")), s)

    matched, name_matched, unmatched = 0, 0, 0
    for i, st in enumerate(parsed["students"]):
        hit, mtype = None, "none"
        if st["reg_no"] and st["reg_no"] in by_reg:
            hit, mtype = by_reg[st["reg_no"]], "reg"
        else:
            hit = by_name.get(_norm_name(st["name"]))
            if hit:
                mtype = "name"
        st["match_type"] = mtype
        st["student_db_id"] = str(hit["_id"]) if hit else ""
        st["matched_name"] = hit.get("full_name", "") if hit else ""
        if not st["reg_no"]:
            st["reg_no"] = (hit or {}).get("student_id", "") or f"TMP-{batch_no or level}-{i + 1}"
        if mtype == "reg":
            matched += 1
        elif mtype == "name":
            name_matched += 1
        else:
            unmatched += 1

    # 🛡 كشف رقم قيد واحد مكتوب لأكثر من طالب داخل الملف (يفسد تجميع السجلات)
    reg_names: dict = {}
    for st in parsed["students"]:
        if st["reg_no"] and not st["reg_no"].startswith("TMP"):
            reg_names.setdefault(st["reg_no"], set()).add(st["name"])
    duplicate_regs = [{"reg_no": k, "names": sorted(v)} for k, v in reg_names.items() if len(v) > 1]

    existing = await db.grade_imports.find_one({
        "department_id": department_id, "level": level,
        "semester_no": semester_no, "academic_year": academic_year,
    })

    result = {
        "courses": parsed["courses"],
        "students": parsed["students"],
        "stats": {"total": len(parsed["students"]), "matched_by_reg": matched,
                  "matched_by_name": name_matched, "unmatched": unmatched},
        "already_imported": bool(existing),
        "duplicate_regs": duplicate_regs,
        "is_template": bool(parsed.get("is_template")),
        "skipped_no_reg": parsed.get("skipped_no_reg", 0),
        "resolved": {"faculty_id": faculty_id, "department_id": department_id,
                     "department_name": resolved.get("department_name", ""),
                     "level": level, "semester_no": semester_no,
                     "academic_year": academic_year, "batch_no": batch_no},
    }
    if commit != "true":
        return result

    # 🛡 منع الاعتماد إذا وُجد رقم قيد واحد لطالبين مختلفين — يجب تصحيح الملف
    if duplicate_regs:
        details = " | ".join(f"{d['reg_no']}: {'، '.join(d['names'])}" for d in duplicate_regs[:5])
        raise HTTPException(status_code=400,
                            detail=f"رقم قيد مكرر لأكثر من طالب في الملف — صحح الملف وأعد الرفع: {details}")

    # 🛡 منع الاعتماد المكرر لنفس (القسم/المستوى/الفصل/العام) — إلا بالاستبدال الصريح
    scope_q = {"department_id": department_id, "level": level,
               "semester_no": semester_no, "academic_year": academic_year}
    if existing and replace != "true":
        raise HTTPException(status_code=409,
                            detail="يوجد استيراد سابق لنفس (القسم/المستوى/الفصل/العام). فعّل خيار «استبدال الاستيراد السابق» أو احذفه من تبويب الاستيرادات أولاً.")
    if existing and replace == "true":
        olds = await db.grade_imports.find(scope_q).to_list(20)
        for o in olds:
            await db.student_grades.delete_many({"import_id": str(o["_id"])})
            await db.grade_imports.delete_one({"_id": o["_id"]})

    imp_doc = {
        "filename": file.filename or "",
        "faculty_id": faculty_id,
        "department_id": department_id,
        "level": level,
        "semester_no": semester_no,
        "academic_year": academic_year,
        "batch_no": batch_no,
        "courses": parsed["courses"],
        "students_count": len(parsed["students"]),
        "imported_by": current_user.get("username", ""),
        "imported_at": datetime.now(timezone.utc).isoformat(),
    }
    ins = await db.grade_imports.insert_one(imp_doc)
    iid = str(ins.inserted_id)
    docs = []
    for st in parsed["students"]:
        docs.append({
            "import_id": iid,
            "student_db_id": st["student_db_id"],
            "student_name": st["name"],
            "reg_no": st["reg_no"],
            "faculty_id": faculty_id,
            "department_id": department_id,
            "level": level,
            "semester_no": semester_no,
            "academic_year": academic_year,
            "batch_no": batch_no,
            "grades": st["grades"],
            "result": st["result"],
            "note": st["note"],
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    if docs:
        await db.student_grades.insert_many(docs)
    await log_activity(current_user, "import_grades", "grades", iid, file.filename or "",
                       {"summary": f"استيراد درجات: مستوى {level} فصل {semester_no} ({academic_year}) — {len(docs)} طالب، {len(parsed['courses'])} مقرر"})
    result["import_id"] = iid
    result["message"] = f"✅ تم اعتماد الاستيراد: {len(docs)} طالب و{len(parsed['courses'])} مقرر"
    return result


@router.get("/grades/imports")
async def list_grade_imports(current_user: dict = Depends(get_current_user)):
    if not _can_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    out = []
    dept_names = {str(d["_id"]): d.get("name", "") async for d in db.departments.find({}, {"name": 1})}
    for i in await db.grade_imports.find().sort("imported_at", -1).to_list(200):
        out.append({
            "id": str(i["_id"]), "filename": i.get("filename", ""),
            "department_id": i.get("department_id", ""),
            "department_name": dept_names.get(i.get("department_id", ""), ""),
            "level": i.get("level"), "semester_no": i.get("semester_no"),
            "academic_year": i.get("academic_year", ""), "batch_no": i.get("batch_no", ""),
            "students_count": i.get("students_count", 0),
            "courses_count": len(i.get("courses", [])),
            "imported_by": i.get("imported_by", ""), "imported_at": (i.get("imported_at") or "")[:16].replace("T", " "),
        })
    return out


@router.delete("/grades/imports/{import_id}")
async def delete_grade_import(import_id: str, current_user: dict = Depends(get_current_user)):
    if not _can_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    imp = await db.grade_imports.find_one({"_id": ObjectId(import_id)})
    if not imp:
        raise HTTPException(status_code=404, detail="الاستيراد غير موجود")
    res = await db.student_grades.delete_many({"import_id": import_id})
    await db.grade_imports.delete_one({"_id": ObjectId(import_id)})
    await log_activity(current_user, "delete_grades_import", "grades", import_id, imp.get("filename", ""),
                       {"summary": f"حذف استيراد درجات «{imp.get('filename', '')}» ({res.deleted_count} سجل)"})
    return {"message": f"تم حذف الاستيراد و{res.deleted_count} سجل درجات"}


# ===== السجل الأكاديمي =====

@router.get("/grades/search")
async def search_grade_students(q: str = "", current_user: dict = Depends(get_current_user)):
    """بحث عن طالب له درجات مستوردة (بالاسم أو رقم القيد)"""
    if not _can_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    query = {}
    if q.strip():
        query = {"$or": [{"student_name": {"$regex": q.strip(), "$options": "i"}},
                         {"reg_no": {"$regex": q.strip(), "$options": "i"}}]}
    groups: dict = {}
    for g in await db.student_grades.find(query).to_list(3000):
        key = g.get("student_db_id") or f"{g.get('reg_no')}|{_norm_name(g.get('student_name', ''))}"
        if key not in groups:
            groups[key] = {"key": key, "student_db_id": g.get("student_db_id", ""),
                           "student_name": g.get("student_name", ""), "reg_no": g.get("reg_no", ""),
                           "semesters": 0}
        groups[key]["semesters"] += 1
    return sorted(groups.values(), key=lambda x: x["student_name"])[:50]


@router.get("/grades/record")
async def get_academic_record(
    student_db_id: str = "",
    reg_no: str = "",
    name: str = "",
    current_user: dict = Depends(get_current_user),
):
    """السجل الأكاديمي الكامل للطالب عبر كل فصوله"""
    if not _can_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    q = {}
    if student_db_id:
        q["student_db_id"] = student_db_id
    elif reg_no:
        q["reg_no"] = reg_no
    elif name:
        q["student_name"] = name
    else:
        raise HTTPException(status_code=400, detail="حدد الطالب")
    recs = await db.student_grades.find(q).to_list(50)
    recs.sort(key=lambda x: (x.get("academic_year", ""), x.get("level") or 0, x.get("semester_no") or 0))
    dept_names = {str(d["_id"]): d.get("name", "") async for d in db.departments.find({}, {"name": 1})}
    out = []
    for rc in recs:
        out.append({
            "id": str(rc["_id"]),
            "student_name": rc.get("student_name", ""),
            "reg_no": rc.get("reg_no", ""),
            "student_db_id": rc.get("student_db_id", ""),
            "department_name": dept_names.get(rc.get("department_id", ""), ""),
            "level": rc.get("level"), "semester_no": rc.get("semester_no"),
            "academic_year": rc.get("academic_year", ""), "batch_no": rc.get("batch_no", ""),
            "grades": rc.get("grades", []), "result": rc.get("result", ""), "note": rc.get("note", ""),
        })
    return out


class GradeRecordUpdate(BaseModel):
    grades: Optional[List[dict]] = None
    result: Optional[str] = None
    note: Optional[str] = None


@router.put("/grades/records/{record_id}")
async def update_grade_record(record_id: str, data: GradeRecordUpdate, current_user: dict = Depends(get_current_user)):
    if not _can_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    rc = await db.student_grades.find_one({"_id": ObjectId(record_id)})
    if not rc:
        raise HTTPException(status_code=404, detail="السجل غير موجود")
    upd = {}
    if data.grades is not None:
        upd["grades"] = data.grades
    if data.result is not None:
        upd["result"] = data.result
    if data.note is not None:
        upd["note"] = data.note
    if upd:
        await db.student_grades.update_one({"_id": ObjectId(record_id)}, {"$set": upd})
        await log_activity(current_user, "update_grade_record", "grades", record_id, rc.get("student_name", ""),
                           {"summary": f"تعديل درجات «{rc.get('student_name', '')}» — مستوى {rc.get('level')} فصل {rc.get('semester_no')}"})
    return {"message": "تم التحديث"}


# ===== بيان حالة ودرجات PDF =====

class GradeStatementRequest(BaseModel):
    record_ids: List[str]
    addressee: str = "إلى من يهمه الأمر"
    status_text: str = ""
    signatory_name: Optional[str] = None
    signatory_title: Optional[str] = None
    base_url: Optional[str] = None


@router.post("/grades/statement")
async def issue_grade_statement(data: GradeStatementRequest, current_user: dict = Depends(get_current_user)):
    """توليد بيان حالة ودرجات رسمي PDF (مع سجل وQR تحقق)"""
    if not _can_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    if not data.record_ids:
        raise HTTPException(status_code=400, detail="اختر فصلاً واحداً على الأقل")
    db = get_db()
    recs = await db.student_grades.find({"_id": {"$in": [ObjectId(r) for r in data.record_ids]}}).to_list(30)
    if not recs:
        raise HTTPException(status_code=404, detail="لا توجد سجلات")
    recs.sort(key=lambda x: (x.get("academic_year", ""), x.get("level") or 0, x.get("semester_no") or 0))
    student_name = recs[0].get("student_name", "")
    reg_no = recs[0].get("reg_no", "")
    faculty_id = recs[0].get("faculty_id", "")
    faculty = None
    try:
        faculty = await db.faculties.find_one({"_id": ObjectId(faculty_id)})
    except Exception:
        pass
    settings = await db.statement_settings.find_one({"_id": f"faculty_{faculty_id}"}) or {}

    token = uuid.uuid4().hex
    verify_url = token
    base = (data.base_url or "").rstrip("/")
    if base:
        verify_url = f"{base}/api/grades/verify/{token}"

    # 🔢 الرقم المرجعي التسلسلي: {تسلسل}/6/2/ت ك ش ق /04/05 (يتغير الرقم الأول فقط)
    from pymongo import ReturnDocument
    cnt = await db.counters.find_one_and_update(
        {"_id": "grade_statement_ref"},
        {"$inc": {"seq": 1}, "$setOnInsert": {"base": 222}},
        upsert=True, return_document=ReturnDocument.AFTER,
    )
    seq_no = cnt.get("base", 222) + cnt.get("seq", 1)
    ref_suffix = settings.get("grade_ref_suffix", " 6 /2  ت ك ش ق /04/05")
    ref_number = f"{seq_no}/{ref_suffix}"

    stmt_doc = {
        "verify_token": token,
        "ref_number": ref_number,
        "student_name": student_name,
        "reg_no": reg_no,
        "faculty_id": faculty_id,
        "faculty_name": (faculty or {}).get("name", ""),
        "addressee": data.addressee,
        "semesters": len(recs),
        "record_ids": data.record_ids,
        "issued_by": current_user.get("username", ""),
        "issued_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.grade_statements.insert_one(stmt_doc)

    pdf = _build_grade_statement_pdf({
        "student_name": student_name,
        "reg_no": reg_no,
        "ref_number": ref_number,
        "faculty_name": (faculty or {}).get("name", "كلية الشريعة والقانون"),
        "addressee": data.addressee,
        "status_text": data.status_text,
        "records": recs,
        "signatory_name": data.signatory_name or settings.get("registrar_name", ""),
        "signatory_title": data.signatory_title or settings.get("signatory_title", "مسجل الكلية"),
        "verify_url": verify_url,
        "issued_at": stmt_doc["issued_at"][:10],
    }, settings)

    await log_activity(current_user, "issue_grade_statement", "grades", token, student_name,
                       {"summary": f"إصدار بيان حالة ودرجات للطالب «{student_name}» ({len(recs)} فصل)"})
    safe_reg = re.sub(r"[^A-Za-z0-9_-]", "", reg_no) or "student"
    return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename=grade_statement_{safe_reg}.pdf"})


@router.get("/grades/verify/{token}")
async def verify_grade_statement(token: str):
    """تحقق عام (بدون تسجيل دخول) من صحة بيان الدرجات"""
    db = get_db()
    s = await db.grade_statements.find_one({"verify_token": token})
    if not s:
        return {"valid": False, "message": "لا يوجد بيان بهذا الرمز — قد تكون الوثيقة غير صحيحة"}
    return {
        "valid": True,
        "message": "بيان حالة ودرجات صحيح صادر رسمياً من جامعة الأحقاف",
        "ref_number": s.get("ref_number", ""),
        "student_name": s.get("student_name", ""),
        "reg_no": s.get("reg_no", ""),
        "faculty_name": s.get("faculty_name", ""),
        "semesters": s.get("semesters", 0),
        "issued_at": (s.get("issued_at") or "")[:10],
    }


def _build_grade_statement_pdf(d: dict, settings: dict) -> bytes:
    import base64
    import textwrap
    import arabic_reshaper
    import qrcode
    from bidi.algorithm import get_display
    from pathlib import Path
    from hijridate import Gregorian
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas as pdfcanvas
    from reportlab.platypus import Table, TableStyle

    def ar(t):
        return get_display(arabic_reshaper.reshape(str(t or "")))

    font_path = Path(__file__).parent.parent / "fonts" / "Amiri-Regular.ttf"
    if "Amiri" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("Amiri", str(font_path)))
    bold_path = Path(__file__).parent.parent / "fonts" / "Amiri-Bold.ttf"
    if bold_path.exists() and "Amiri-Bold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("Amiri-Bold", str(bold_path)))
    BOLD = "Amiri-Bold" if "Amiri-Bold" in pdfmetrics.getRegisteredFontNames() else "Amiri"

    buf = io.BytesIO()
    W, H = A4
    c = pdfcanvas.Canvas(buf, pagesize=A4)

    def draw_header():
        logo_b64 = settings.get("logo_base64") or ""
        img = None
        if logo_b64:
            try:
                img = ImageReader(io.BytesIO(base64.b64decode(logo_b64.split(",")[-1])))
            except Exception:
                img = None
        if img is None:
            default_logo = Path(__file__).parent.parent / "assets" / "university_logo.jpeg"
            if default_logo.exists():
                img = ImageReader(str(default_logo))
        if img:
            c.drawImage(img, W / 2 - 13 * mm, H - 36 * mm, 26 * mm, 26 * mm, mask="auto", preserveAspectRatio=True)
        c.setFont("Amiri", 15)
        c.drawRightString(W - 18 * mm, H - 17 * mm, ar("جامعة الأحقاف"))
        c.setFont("Amiri", 12)
        c.drawRightString(W - 18 * mm, H - 24 * mm, ar(d.get("faculty_name", "")))
        c.drawRightString(W - 18 * mm, H - 31 * mm, ar("مكتب القبول والتسجيل"))
        c.setFont("Helvetica-Bold", 11)
        c.drawString(18 * mm, H - 17 * mm, "AL-AHGAFF UNIVERSITY")
        c.setFont("Helvetica", 9)
        c.drawString(18 * mm, H - 24 * mm, settings.get("faculty_name_en", ""))
        c.setLineWidth(1.3)
        c.line(18 * mm, H - 39 * mm, W - 18 * mm, H - 39 * mm)
        c.setLineWidth(0.4)
        c.line(18 * mm, H - 40.4 * mm, W - 18 * mm, H - 40.4 * mm)

    draw_header()

    issued = d.get("issued_at", "")
    try:
        y0, m0, d0 = [int(x) for x in issued.split("-")]
        hj = Gregorian(y0, m0, d0).to_hijri()
        hijri_str = f"{hj.year}/{hj.month:02d}/{hj.day:02d}هـ"
    except Exception:
        hijri_str = ""
    c.setFont(BOLD, 11)
    c.drawRightString(W - 18 * mm, H - 47 * mm, ar(f"التاريخ: {hijri_str}"))
    c.drawRightString(W - 18 * mm, H - 53 * mm, ar(f"الموافق: {issued.replace('-', '/')}م"))
    # 🔢 الرقم المرجعي بالأخضر — يسار الترويسة مقابل التاريخ
    if d.get("ref_number"):
        c.setFillColorRGB(0.0, 0.45, 0.13)
        c.drawString(18 * mm, H - 47 * mm, ar(f"الـرقـــم : {d['ref_number']}"))
        c.setFillColorRGB(0, 0, 0)

    yy = H - 62 * mm
    c.setFont(BOLD, 13)
    c.drawRightString(W - 18 * mm, yy, ar(f"الفاضل / {d.get('addressee', '')}"))
    c.drawString(18 * mm, yy, ar("المحترم"))
    yy -= 8 * mm
    c.setFont("Amiri", 12.5)
    c.drawCentredString(W / 2, yy, ar("السلام عليكم ورحمة الله وبركاته،،،"))
    yy -= 9 * mm
    c.setFont(BOLD, 13)
    c.drawCentredString(W / 2, yy, ar(f"الموضوع / بيان حالة ودرجات الطالب: {d.get('student_name', '')}"))
    if d.get("reg_no"):
        yy -= 7 * mm
        c.setFont("Amiri", 12)
        c.drawCentredString(W / 2, yy, ar(f"رقم القيد: ({d['reg_no']})"))
    yy -= 9 * mm
    c.setFont("Amiri", 12)
    for para in (d.get("status_text") or "").split("\n"):
        for line in (textwrap.wrap(para, width=88) or [""]):
            c.drawRightString(W - 18 * mm, yy, ar(line))
            yy -= 7 * mm
    yy -= 3 * mm

    # ===== جداول الفصول (المقررات أعمدة كما في النموذج) =====
    for rc in d.get("records", []):
        grades = [g for g in rc.get("grades", []) if str(g.get("course_name", "")).strip()]
        if not grades:
            continue
        n = len(grades)
        label = f"الفصل {TERM_AR.get(rc.get('semester_no'), rc.get('semester_no'))} — المستوى {LEVEL_AR.get(rc.get('level'), rc.get('level'))} — {rc.get('academic_year', '')}"
        if rc.get("result"):
            label += f"   |   النتيجة: {rc['result']}"
        needed = 14 * mm + 34 * mm
        if yy - needed < 40 * mm:
            c.showPage()
            draw_header()
            yy = H - 50 * mm
        c.setFont(BOLD, 11.5)
        c.setFillColorRGB(0.05, 0.25, 0.1)
        c.drawRightString(W - 18 * mm, yy, ar(label))
        c.setFillColorRGB(0, 0, 0)
        yy -= 6 * mm

        # صفوف: أسماء المقررات / الساعات / الدرجة — والمقررات أعمدة (معكوسة للاتجاه العربي)
        rev = list(reversed(grades))

        def wrap_ar(t, wchars):
            lines = textwrap.wrap(str(t), width=wchars) or [""]
            return "\n".join(ar(x) for x in lines)

        avail = W - 36 * mm - 26 * mm
        col_w = max(avail / n, 16 * mm)
        wchars = max(int(col_w / (2.1 * mm)), 6)
        row0 = [wrap_ar(g["course_name"], wchars) for g in rev] + [ar("المقرر")]
        row1 = [ar(_fmt_val(g.get("credits", ""))) for g in rev] + [ar("الساعات")]
        row2 = [ar(_fmt_val(g.get("total", ""))) for g in rev] + [ar("الدرجة")]
        rows = [row0, row1, row2]
        heights = [16 * mm, 7 * mm, 8 * mm]
        if any(g.get("grade_letter") for g in grades):
            rows.append([ar(_fmt_val(g.get("grade_letter", ""))) for g in rev] + [ar("التقدير")])
            heights.append(7 * mm)
        tbl = Table(rows, colWidths=[col_w] * n + [26 * mm],
                    rowHeights=heights)
        tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Amiri"),
            ("FONTNAME", (-1, 0), (-1, -1), BOLD),
            ("FONTNAME", (0, 2), (-1, 2), BOLD),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
            ("BACKGROUND", (-1, 0), (-1, -1), colors.Color(0.92, 0.95, 0.92)),
            ("BACKGROUND", (0, 0), (-2, 0), colors.Color(0.96, 0.96, 0.9)),
        ]))
        tw, th = tbl.wrapOn(c, avail + 26 * mm, 40 * mm)
        tbl.drawOn(c, W - 18 * mm - tw, yy - th)
        yy -= th + 4 * mm
        if rc.get("note"):
            c.setFont("Amiri", 9.5)
            c.setFillColorRGB(0.4, 0.25, 0)
            c.drawRightString(W - 18 * mm, yy, ar(f"ملاحظة: {rc['note']}"))
            c.setFillColorRGB(0, 0, 0)
            yy -= 6 * mm
        yy -= 3 * mm

    # ===== التوقيع + QR =====
    if yy < 55 * mm:
        c.showPage()
        draw_header()
        yy = H - 55 * mm
    sig_img = None
    sig_b64 = settings.get("signature_base64") or ""
    if sig_b64:
        try:
            sig_img = ImageReader(io.BytesIO(base64.b64decode(sig_b64.split(",")[-1])))
        except Exception:
            sig_img = None
    c.setFont("Amiri", 13)
    c.drawString(28 * mm, yy - 10 * mm, ar(d.get("signatory_title", "مسجل الكلية")))
    name_y = yy - 18 * mm
    if sig_img:
        c.drawImage(sig_img, 20 * mm, yy - 28 * mm, 36 * mm, 14 * mm, mask="auto", preserveAspectRatio=True)
        name_y = yy - 32 * mm
    c.setFont(BOLD, 12)
    c.drawString(24 * mm, name_y, ar(d.get("signatory_name", "")))

    qr_img = qrcode.make(d.get("verify_url", ""), box_size=4, border=1)
    qb = io.BytesIO()
    qr_img.save(qb, format="PNG")
    qb.seek(0)
    c.drawImage(ImageReader(qb), W - 46 * mm, 28 * mm, 24 * mm, 24 * mm)
    c.setFont("Helvetica", 7.5)
    c.drawCentredString(W - 34 * mm, 24.5 * mm, "Scan to verify")

    c.line(18 * mm, 20 * mm, W - 18 * mm, 20 * mm)
    c.setFont("Amiri", 9)
    footer_parts = []
    for k, pre in (("address", ""), ("phones", "تلفون: "), ("po_box", "ص.ب "), ("website", "")):
        if settings.get(k):
            footer_parts.append(f"{pre}{settings[k]}")
    c.drawCentredString(W / 2, 14 * mm, ar(" — ".join(footer_parts)))

    c.showPage()
    c.save()
    return buf.getvalue()


# ===== 📤 تصدير تحليل النتيجة (PDF / Excel) لمجلس الكلية =====

_STAGE_LABELS = {"saai": "السعي فقط (إنذار مبكر)", "round1": "بعد الدور الأول (نتائج أولية)", "final": "النتيجة المكتملة"}


def _analysis_fname(a: dict, ext: str) -> str:
    info = a.get("info", {})
    parts = ["تحليل النتيجة"]
    if info.get("level"):
        parts.append(f"م{info['level']}")
    if info.get("semester_no"):
        parts.append(f"فصل {info['semester_no']}")
    if info.get("academic_year"):
        parts.append(str(info["academic_year"]))
    parts.append(datetime.now(timezone.utc).strftime("%Y-%m-%d"))
    return " - ".join(parts) + f".{ext}"


def _analysis_meta_rows(a: dict):
    info = a.get("info", {})
    rows = [["الكشف", info.get("filename", "")],
            ["المستوى", info.get("level", "")],
            ["الفصل", info.get("semester_no", "")],
            ["العام الجامعي", info.get("academic_year", "")],
            ["الدفعة", info.get("batch_no", "")],
            ["عدد الطلاب", info.get("students", 0)],
            ["مرحلة التحليل", _STAGE_LABELS.get(a.get("stage", ""), a.get("stage", ""))]]
    if a.get("stage") != "saai":
        sm = a.get("summary", {})
        rows += [["ناجح بكل المواد", sm.get("passed", 0)],
                 ["لديهم مواد راسبة (دور ثانٍ)", sm.get("second_round", 0)],
                 ["مفصولون", sm.get("dismissed", 0)]]
    return rows


def _analysis_course_table(a: dict):
    saai = a.get("stage") == "saai"
    if saai:
        head = ["المقرر", "الساعات", "العدد", "متوسط السعي", "الأعلى", "الأدنى", "سعي ضعيف", "غياب"]
        rows = [[c["course"], c.get("credits", ""), c["count"], c["avg"], c["max"], c["min"],
                 c.get("weak", 0), c.get("absent", 0)] for c in a.get("course_stats", [])]
    else:
        head = ["المقرر", "الساعات", "العدد", "المتوسط", "الأعلى", "الأدنى", "راسبون", "نسبة النجاح %", "متوسط السعي", "غياب"]
        rows = [[c["course"], c.get("credits", ""), c["count"], c["avg"], c["max"], c["min"],
                 c.get("fail_count", 0), c.get("pass_rate", ""),
                 c.get("avg_saai", "") if c.get("avg_saai") is not None else "", c.get("absent", 0)]
                for c in a.get("course_stats", [])]
    return head, rows


@router.get("/grades/analysis/{import_id}/export/excel")
async def export_analysis_excel(import_id: str, current_user: dict = Depends(get_current_user)):
    if not _can_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    a = await _compute_analysis(get_db(), import_id)
    from urllib.parse import quote
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet(title, head, rows):
        ws = wb.create_sheet(title=title[:31])
        ws.sheet_view.rightToLeft = True
        ws.append([f"تقرير تحليل النتيجة — {title}"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(head), 2))
        t = ws.cell(row=1, column=1)
        t.font = Font(bold=True, size=13, color="1A237E")
        t.alignment = Alignment(horizontal="center")
        ws.append(head)
        for c in ws[2]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="3949AB")
            c.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append(r)
        for col in ws.iter_cols(min_row=2):
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(w + 4, 10), 50)
        return ws

    add_sheet("الملخص", ["البيان", "القيمة"], _analysis_meta_rows(a))
    head, rows = _analysis_course_table(a)
    add_sheet("إحصاءات المقررات", head, rows)
    if a.get("top"):
        add_sheet("الأوائل", ["م", "الاسم", "رقم القيد", "المعدل"],
                  [[i + 1, t.get("name", ""), t.get("reg_no", ""), t.get("avg", "")] for i, t in enumerate(a["top"])])
    if a.get("second_round"):
        add_sheet("الدور الثاني", ["الاسم", "رقم القيد", "عدد المواد", "المواد الراسب فيها"],
                  [[x.get("name", ""), x.get("reg_no", ""), len(x.get("courses", [])), "، ".join(x.get("courses", []))] for x in a["second_round"]])
    if a.get("fail_lists") and any(v for v in a["fail_lists"].values()):
        add_sheet("الراسبون بالمقرر", ["المقرر", "العدد", "الأسماء"],
                  [[cn, len(names), "، ".join(names)] for cn, names in a["fail_lists"].items() if names])
    if a.get("saai_alerts"):
        add_sheet("إنذار السعي", ["الاسم", "رقم القيد", "المقررات (سعي < النصف)"],
                  [[x.get("name", ""), x.get("reg_no", ""), "، ".join(x.get("courses", []))] for x in a["saai_alerts"]])
    if a.get("dismissed"):
        add_sheet("المفصولون", ["الاسم", "الملاحظة"],
                  [[x.get("name", ""), x.get("note", "")] for x in a["dismissed"]])

    out = io.BytesIO()
    wb.save(out)
    fname = quote(_analysis_fname(a, "xlsx"))
    return StreamingResponse(io.BytesIO(out.getvalue()),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}",
                                      "X-Filename": fname})


@router.get("/grades/analysis/{import_id}/export/pdf")
async def export_analysis_pdf(import_id: str, current_user: dict = Depends(get_current_user)):
    if not _can_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    a = await _compute_analysis(get_db(), import_id)
    from urllib.parse import quote
    import arabic_reshaper
    from bidi.algorithm import get_display
    from pathlib import Path
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER

    def ar(t):
        return get_display(arabic_reshaper.reshape(str(t if t is not None else "")))

    fonts_dir = Path(__file__).parent.parent / "fonts"
    if "Amiri" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("Amiri", str(fonts_dir / "Amiri-Regular.ttf")))
    if (fonts_dir / "Amiri-Bold.ttf").exists() and "Amiri-Bold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("Amiri-Bold", str(fonts_dir / "Amiri-Bold.ttf")))
    BOLD = "Amiri-Bold" if "Amiri-Bold" in pdfmetrics.getRegisteredFontNames() else "Amiri"

    title_st = ParagraphStyle("t", fontName=BOLD, fontSize=16, alignment=TA_CENTER, spaceAfter=4)
    sub_st = ParagraphStyle("s", fontName="Amiri", fontSize=11, alignment=TA_CENTER, textColor=colors.HexColor("#455a64"))
    sec_st = ParagraphStyle("h", fontName=BOLD, fontSize=12.5, alignment=TA_RIGHT, textColor=colors.HexColor("#1a237e"), spaceBefore=10, spaceAfter=4)

    def rtl_table(head, rows, col_widths=None, highlight_col=None):
        data = [[ar(h) for h in reversed(head)]] + [[ar(v) for v in reversed(r)] for r in rows]
        t = Table(data, colWidths=list(reversed(col_widths)) if col_widths else None, repeatRows=1)
        style = [("FONTNAME", (0, 0), (-1, -1), "Amiri"),
                 ("FONTNAME", (0, 0), (-1, 0), BOLD),
                 ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                 ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3949ab")),
                 ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                 ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                 ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                 ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#b0bec5")),
                 ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f7ff")]),
                 ("TOPPADDING", (0, 0), (-1, -1), 3),
                 ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]
        t.setStyle(TableStyle(style))
        return t

    info = a.get("info", {})
    story = [
        Paragraph(ar("جامعة الأحقاف"), title_st),
        Paragraph(ar("تقرير تحليل النتيجة"), title_st),
        Paragraph(ar(f"{info.get('filename', '')} — المستوى {info.get('level', '')} · الفصل {info.get('semester_no', '')} · {info.get('academic_year', '')}"), sub_st),
        Paragraph(ar(f"مرحلة التحليل: {_STAGE_LABELS.get(a.get('stage', ''), '')} — عدد الطلاب: {info.get('students', 0)}"), sub_st),
        Spacer(1, 6 * mm),
    ]

    if a.get("stage") != "saai":
        sm = a.get("summary", {})
        story.append(rtl_table(["ناجح بكل المواد", "لديهم مواد راسبة (دور ثانٍ)", "مفصولون"],
                               [[sm.get("passed", 0), sm.get("second_round", 0), sm.get("dismissed", 0)]],
                               col_widths=[85 * mm, 85 * mm, 85 * mm]))
        story.append(Spacer(1, 4 * mm))

    story.append(Paragraph(ar("إحصاءات المقررات"), sec_st))
    head, rows = _analysis_course_table(a)
    story.append(rtl_table(head, rows))

    if a.get("top"):
        story.append(Paragraph(ar("العشرة الأوائل"), sec_st))
        story.append(rtl_table(["م", "الاسم", "رقم القيد", "المعدل"],
                               [[i + 1, t.get("name", ""), t.get("reg_no", ""), t.get("avg", "")] for i, t in enumerate(a["top"])]))

    if a.get("second_round"):
        story.append(Paragraph(ar(f"طلاب الدور الثاني ({len(a['second_round'])})"), sec_st))
        story.append(rtl_table(["الاسم", "رقم القيد", "المواد الراسب فيها"],
                               [[x.get("name", ""), x.get("reg_no", ""), "، ".join(x.get("courses", []))] for x in a["second_round"]]))

    if a.get("saai_alerts"):
        story.append(Paragraph(ar(f"إنذار مبكر — سعي أقل من النصف ({len(a['saai_alerts'])})"), sec_st))
        story.append(rtl_table(["الاسم", "رقم القيد", "المقررات"],
                               [[x.get("name", ""), x.get("reg_no", ""), "، ".join(x.get("courses", []))] for x in a["saai_alerts"]]))

    if a.get("dismissed"):
        story.append(Paragraph(ar(f"المفصولون ({len(a['dismissed'])})"), sec_st))
        story.append(rtl_table(["الاسم", "الملاحظة"],
                               [[x.get("name", ""), x.get("note", "")] for x in a["dismissed"]]))

    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph(ar(f"أُصدر آلياً من نظام إدارة الجامعة — {datetime.now(timezone.utc).strftime('%Y-%m-%d')}"),
                           ParagraphStyle("f", fontName="Amiri", fontSize=8.5, alignment=TA_CENTER, textColor=colors.HexColor("#90a4ae"))))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    doc.build(story)
    fname = quote(_analysis_fname(a, "pdf"))
    return StreamingResponse(io.BytesIO(buf.getvalue()), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}",
                                      "X-Filename": fname})
