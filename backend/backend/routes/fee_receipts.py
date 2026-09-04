"""💰 السندات المالية: رفع الطلاب لسندات الرسوم (تجديد القيد/القسم الداخلي/أخرى) وتعميدها"""
import logging
from datetime import datetime, timezone, timedelta
from typing import Optional
from fastapi import APIRouter, HTTPException, Depends
from pydantic import BaseModel
from bson import ObjectId

from .deps import get_db, get_current_user, has_permission, log_activity

router = APIRouter(tags=["السندات المالية"])

BUILTIN_TYPES = [
    {"key": "renewal", "name": "تجديد القيد الدراسي", "builtin": True},
]
STATUS_LABELS = {"pending": "قيد المراجعة", "approved": "مقبول", "rejected": "مرفوض", "not_paid": "غير دافع"}


def _now():
    return datetime.now(timezone(timedelta(hours=3))).isoformat()


def can_manage_fees(user: dict) -> bool:
    return user.get("role") == "admin" or has_permission(user, "manage_fee_receipts")


async def _seed_types(db):
    for t in BUILTIN_TYPES:
        await db.fee_types.update_one({"key": t["key"]}, {"$setOnInsert": {**t, "is_active": True, "created_at": _now()}}, upsert=True)
    # هجرة لمرة واحدة: القسم الداخلي لم يعد نوعاً محمياً — يتحول لشهري قابل للحذف
    await db.fee_types.update_many({"key": "dormitory", "builtin": True},
                                   {"$set": {"builtin": False, "recurring": True}})


async def _academic_year(db) -> str:
    sem = await db.semesters.find_one({"status": "active"})
    return (sem or {}).get("academic_year", "") or ""


def _date_warning(receipt_date: str, year: str) -> bool:
    """⚠️ تاريخ السند خارج العام الجامعي الحالي (مثال year: 2024-2025)"""
    try:
        y = int((receipt_date or "")[:4])
        years = [int(p) for p in (year or "").split("-") if p.strip().isdigit()]
        return bool(years) and y not in years
    except Exception:
        return False


async def _notify_student(db, student: dict, title: str, message: str):
    try:
        uid = student.get("user_id")
        if not uid:
            return
        await db.notifications.insert_one({"user_id": uid, "title": title, "message": message,
                                           "type": "general", "is_read": False, "created_at": _now()})
        from services.firebase_service import send_notification_to_many
        tokens = [d["token"] for d in await db.fcm_tokens.find({"user_id": uid}).to_list(20) if d.get("token")]
        if tokens:
            await send_notification_to_many(tokens, title, message)
    except Exception as e:
        logging.error(f"fee notify error: {e}")


# ===== أنواع الرسوم (ديناميكية) =====

class FeeTypeCreate(BaseModel):
    name: str
    recurring: bool = False


class FeeTypeUpdate(BaseModel):
    recurring: bool


@router.get("/fees/types")
async def list_fee_types(current_user: dict = Depends(get_current_user)):
    db = get_db()
    await _seed_types(db)
    types = [{"id": str(t["_id"]), "key": t.get("key", ""), "name": t["name"], "builtin": bool(t.get("builtin")),
              "recurring": bool(t.get("recurring"))}
             for t in await db.fee_types.find({"is_active": {"$ne": False}}).to_list(100)]
    types.append({"id": "other", "key": "other", "name": "أخرى (نوع حر)", "builtin": True, "recurring": False})
    return {"types": types, "academic_year": await _academic_year(db)}


@router.put("/fees/types/{type_id}")
async def update_fee_type(type_id: str, data: FeeTypeUpdate, current_user: dict = Depends(get_current_user)):
    """تبديل نوع الرسوم بين سنوي ومتكرر (شهري)"""
    if not can_manage_fees(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    t = await db.fee_types.find_one({"_id": ObjectId(type_id)})
    if not t:
        raise HTTPException(status_code=404, detail="غير موجود")
    await db.fee_types.update_one({"_id": t["_id"]}, {"$set": {"recurring": data.recurring}})
    await log_activity(current_user, "update_fee_type", "fee_type", type_id, t.get("name"),
                       {"summary": f"«{t.get('name')}» أصبح {'متكرراً (شهرياً)' if data.recurring else 'سنوياً'}"})
    return {"message": "تم التحديث"}


@router.post("/fees/types")
async def create_fee_type(data: FeeTypeCreate, current_user: dict = Depends(get_current_user)):
    if not can_manage_fees(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    name = data.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="الاسم مطلوب")
    if await db.fee_types.find_one({"name": name, "is_active": {"$ne": False}}):
        raise HTTPException(status_code=400, detail="النوع موجود مسبقاً")
    r = await db.fee_types.insert_one({"name": name, "builtin": False, "recurring": bool(data.recurring),
                                       "is_active": True, "created_at": _now()})
    await log_activity(current_user, "create_fee_type", "fee_type", str(r.inserted_id), name,
                       {"summary": f"إضافة نوع رسوم «{name}»" + (" (متكرر/شهري)" if data.recurring else "")})
    return {"id": str(r.inserted_id), "message": "تمت الإضافة"}


@router.delete("/fees/types/{type_id}")
async def delete_fee_type(type_id: str, current_user: dict = Depends(get_current_user)):
    if not can_manage_fees(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    t = await db.fee_types.find_one({"_id": ObjectId(type_id)})
    if not t:
        raise HTTPException(status_code=404, detail="غير موجود")
    if t.get("builtin"):
        raise HTTPException(status_code=400, detail="نوع أساسي لا يمكن حذفه")
    await db.fee_types.update_one({"_id": t["_id"]}, {"$set": {"is_active": False}})
    return {"message": "تم الحذف"}


# ===== الطالب: رفع سند + حالتي =====

class ReceiptUpload(BaseModel):
    type_id: str  # id من fee_types أو "other"
    other_label: Optional[str] = ""
    image_base64: str
    receipt_no: Optional[str] = ""
    amount: Optional[str] = ""
    receipt_date: Optional[str] = ""  # 📅 تاريخ السند الورقي YYYY-MM-DD
    statement: Optional[str] = ""  # 📝 بيان الدفعة (إلزامي للرسوم المتكررة: «تغذية شهر يناير»)
    notes: Optional[str] = ""


async def _get_student(db, current_user):
    if current_user.get("role") != "student":
        raise HTTPException(status_code=403, detail="هذه الخدمة للطلاب فقط")
    student = await db.students.find_one({"user_id": current_user["id"]})
    if not student:
        raise HTTPException(status_code=404, detail="سجل الطالب غير موجود")
    return student


async def _type_info(db, type_id: str, other_label: str = "") -> tuple:
    """(اسم النوع، هل متكرر)"""
    if type_id == "other":
        return ((other_label or "").strip() or "رسوم أخرى", False)
    t = await db.fee_types.find_one({"_id": ObjectId(type_id)})
    if not t:
        raise HTTPException(status_code=400, detail="نوع الرسوم غير موجود")
    return (t["name"], bool(t.get("recurring")))


async def _type_name(db, type_id: str, other_label: str = "") -> str:
    return (await _type_info(db, type_id, other_label))[0]


@router.post("/fees/receipts")
async def upload_receipt(data: ReceiptUpload, current_user: dict = Depends(get_current_user)):
    db = get_db()
    student = await _get_student(db, current_user)
    if not data.image_base64 or len(data.image_base64) < 100:
        raise HTTPException(status_code=400, detail="صورة السند مطلوبة")
    if len(data.image_base64) > 3_000_000:
        raise HTTPException(status_code=400, detail="حجم الصورة كبير — الحد الأقصى 2MB")
    if data.type_id == "other" and not (data.other_label or "").strip():
        raise HTTPException(status_code=400, detail="اكتب نوع الرسوم في الحقل الحر")
    year = await _academic_year(db)
    type_name, recurring = await _type_info(db, data.type_id, data.other_label)
    statement = (data.statement or "").strip()
    if recurring and not statement:
        raise HTTPException(status_code=400, detail=f"«{type_name}» رسوم متكررة — اكتب بيان الدفعة (مثال: {type_name} شهر يناير)")
    sid = str(student["_id"])
    match: dict = {"student_id": sid, "type_id": data.type_id, "academic_year": year}
    if data.type_id == "other":
        match["type_name"] = type_name
    if recurring:
        match["statement"] = statement  # كل دفعة ببيانها المستقل
    blocking = await db.fee_receipts.find_one({**match, "status": {"$in": ["pending", "approved"]}})
    if blocking:
        if blocking["status"] == "approved":
            raise HTTPException(status_code=400,
                                detail=f"دفعة «{statement}» معتمدة مسبقاً" if recurring else "سندك لهذا النوع معتمد مسبقاً")
        raise HTTPException(status_code=400,
                            detail=f"دفعة «{statement}» قيد المراجعة حالياً — لا يمكن رفع سند جديد" if recurring
                            else "لديك سند لهذا النوع قيد المراجعة حالياً — لا يمكن رفع سند جديد")
    existing = await db.fee_receipts.find_one({**match, "status": "rejected"})
    doc = {
        "student_id": sid, "type_id": data.type_id, "type_name": type_name,
        "academic_year": year, "image_base64": data.image_base64,
        "receipt_no": (data.receipt_no or "").strip(), "amount": (data.amount or "").strip(),
        "receipt_date": (data.receipt_date or "").strip(),
        "statement": statement, "recurring": recurring,
        "notes": (data.notes or "").strip(), "status": "pending",
        "rejection_reason": "", "uploaded_at": _now(), "reviewed_by": "", "reviewed_at": "",
    }
    if existing:
        await db.fee_receipts.update_one({"_id": existing["_id"]}, {"$set": doc})
        rid = str(existing["_id"])
    else:
        rid = str((await db.fee_receipts.insert_one(doc)).inserted_id)
    return {"id": rid, "message": "تم رفع السند وهو الآن قيد المراجعة", "status": "pending"}


@router.get("/fees/my-status")
async def my_fee_status(current_user: dict = Depends(get_current_user)):
    db = get_db()
    student = await _get_student(db, current_user)
    await _seed_types(db)
    year = await _academic_year(db)
    receipts = await db.fee_receipts.find({"student_id": str(student["_id"]), "academic_year": year},
                                          {"image_base64": 0}).to_list(50)
    by_type = {}
    for r in receipts:
        by_type.setdefault(r["type_id"], []).append(r)
    out = []
    async for t in db.fee_types.find({"is_active": {"$ne": False}}):
        tid = str(t["_id"])
        rs = sorted(by_type.get(tid, []), key=lambda x: x.get("uploaded_at", ""), reverse=True)
        recurring = bool(t.get("recurring"))
        latest = rs[0] if rs else None
        status = latest["status"] if latest else "not_paid"
        approved_count = sum(1 for r in rs if r["status"] == "approved")
        if recurring and approved_count > 0 and status != "pending":
            status = "approved"
        out.append({"type_id": tid, "type_name": t["name"], "status": status,
                    "status_label": (f"دافع ({approved_count} دفعات)" if recurring and approved_count else STATUS_LABELS.get(status, status)),
                    "recurring": recurring, "paid_count": approved_count,
                    "receipt_id": str(latest["_id"]) if latest else None,
                    "payments": [{"receipt_id": str(r["_id"]),
                                  "statement": r.get("statement", ""), "status": r["status"],
                                  "status_label": STATUS_LABELS.get(r["status"], ""),
                                  "rejection_reason": r.get("rejection_reason", ""),
                                  "uploaded_at": r.get("uploaded_at", "")} for r in rs] if recurring else [],
                    "rejection_reason": (latest or {}).get("rejection_reason", "")})
    others = [{"type_id": "other", "type_name": r["type_name"], "status": r["status"],
               "status_label": STATUS_LABELS.get(r["status"], r["status"]),
               "recurring": False, "paid_count": 1 if r["status"] == "approved" else 0, "payments": [],
               "receipt_id": str(r["_id"]),
               "statement": r.get("statement", ""),
               "rejection_reason": r.get("rejection_reason", "")}
              for r in receipts if r["type_id"] == "other"]
    return {"academic_year": year, "statuses": out + others}


# ===== الإدارة: المراجعة والتعميد =====

@router.get("/fees/receipts")
async def list_receipts(status: Optional[str] = None, type_id: Optional[str] = None,
                        department_id: Optional[str] = None,
                        current_user: dict = Depends(get_current_user)):
    if not can_manage_fees(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    year = await _academic_year(db)
    q: dict = {"academic_year": year}
    if status:
        q["status"] = status
    if type_id:
        q["type_id"] = type_id
    receipts = await db.fee_receipts.find(q, {"image_base64": 0}).sort("uploaded_at", -1).to_list(1000)
    sids = [ObjectId(r["student_id"]) for r in receipts if ObjectId.is_valid(r["student_id"])]
    smap = {str(s["_id"]): s for s in await db.students.find({"_id": {"$in": sids}}).to_list(2000)}
    deps = {str(d["_id"]): d.get("name", "") for d in await db.departments.find({}, {"name": 1}).to_list(500)}
    # كشف تكرار رقم السند (عبر كل سندات العام بغض النظر عن الفلتر)
    dup_nos = {}
    async for rr in db.fee_receipts.find({"academic_year": year, "receipt_no": {"$exists": True, "$ne": ""}}, {"receipt_no": 1}):
        no = rr.get("receipt_no", "")
        if no:
            dup_nos[no] = dup_nos.get(no, 0) + 1
    out = []
    for r in receipts:
        s = smap.get(r["student_id"], {})
        if department_id and s.get("department_id") != department_id:
            continue
        out.append({
            "id": str(r["_id"]), "student_id": r["student_id"],
            "student_name": s.get("full_name", "؟"), "student_number": s.get("student_id", ""),
            "department_name": deps.get(s.get("department_id", ""), ""), "level": s.get("level"),
            "type_name": r["type_name"], "receipt_no": r.get("receipt_no", ""), "amount": r.get("amount", ""),
            "statement": r.get("statement", ""),
            "receipt_date": r.get("receipt_date", ""),
            "date_warning": _date_warning(r.get("receipt_date", ""), year),
            "notes": r.get("notes", ""), "status": r["status"], "status_label": STATUS_LABELS.get(r["status"], ""),
            "rejection_reason": r.get("rejection_reason", ""), "uploaded_at": r.get("uploaded_at", ""),
            "reviewed_by": r.get("reviewed_by", ""), "reviewed_at": r.get("reviewed_at", ""),
            "duplicate_receipt_no": bool(r.get("receipt_no")) and dup_nos.get(r.get("receipt_no"), 0) > 1,
            "manual_entry": bool(r.get("manual_entry")),
        })
    return {"receipts": out, "academic_year": year}


@router.get("/fees/receipts/{receipt_id}/image")
async def receipt_image(receipt_id: str, current_user: dict = Depends(get_current_user)):
    db = get_db()
    r = await db.fee_receipts.find_one({"_id": ObjectId(receipt_id)})
    if not r:
        raise HTTPException(status_code=404, detail="غير موجود")
    if not can_manage_fees(current_user):
        student = await db.students.find_one({"user_id": current_user["id"]})
        if not student or str(student["_id"]) != r["student_id"]:
            raise HTTPException(status_code=403, detail="غير مصرح لك")
    return {"image_base64": r.get("image_base64", "")}


class RejectBody(BaseModel):
    reason: str


async def _review(db, receipt_id, current_user, status, reason=""):
    if not can_manage_fees(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    r = await db.fee_receipts.find_one({"_id": ObjectId(receipt_id)})
    if not r:
        raise HTTPException(status_code=404, detail="السند غير موجود")
    await db.fee_receipts.update_one({"_id": r["_id"]}, {"$set": {
        "status": status, "rejection_reason": reason,
        "reviewed_by": current_user.get("full_name") or current_user.get("username", ""),
        "reviewed_at": _now()}})
    student = await db.students.find_one({"_id": ObjectId(r["student_id"])}) if ObjectId.is_valid(r["student_id"]) else None
    if student:
        if status == "approved":
            await _notify_student(db, student, f"✅ تم اعتماد سند {r['type_name']}",
                                  f"تم اعتماد سند «{r['type_name']}» — أصبحت حالتك: دافع.")
        else:
            await _notify_student(db, student, f"❌ رُفض سند {r['type_name']}",
                                  f"رُفض سندك لـ«{r['type_name']}». السبب: {reason}. يرجى إعادة رفع سند صحيح.")
    await log_activity(current_user, f"fee_receipt_{status}", "fee_receipt", receipt_id,
                       (student or {}).get("full_name", ""),
                       {"summary": f"{'اعتماد' if status == 'approved' else 'رفض'} سند «{r['type_name']}» للطالب {(student or {}).get('full_name', '؟')}" + (f" — السبب: {reason}" if reason else "")})
    return {"message": "تم الاعتماد ✅" if status == "approved" else "تم الرفض وإشعار الطالب"}


@router.post("/fees/receipts/{receipt_id}/approve")
async def approve_receipt(receipt_id: str, current_user: dict = Depends(get_current_user)):
    return await _review(get_db(), receipt_id, current_user, "approved")


@router.post("/fees/receipts/{receipt_id}/reject")
async def reject_receipt(receipt_id: str, data: RejectBody, current_user: dict = Depends(get_current_user)):
    if not (data.reason or "").strip():
        raise HTTPException(status_code=400, detail="سبب الرفض مطلوب")
    return await _review(get_db(), receipt_id, current_user, "rejected", data.reason.strip())


@router.get("/fees/stats")
async def fee_stats(current_user: dict = Depends(get_current_user)):
    if not can_manage_fees(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    await _seed_types(db)
    year = await _academic_year(db)
    total_students = await db.students.count_documents({"is_active": True})
    out = []
    async for t in db.fee_types.find({"is_active": {"$ne": False}}):
        tid = str(t["_id"])
        base = {"type_id": tid, "academic_year": year}
        approved = await db.fee_receipts.count_documents({**base, "status": "approved"})
        pending = await db.fee_receipts.count_documents({**base, "status": "pending"})
        rejected = await db.fee_receipts.count_documents({**base, "status": "rejected"})
        recurring = bool(t.get("recurring"))
        if recurring:
            paid_students = len(await db.fee_receipts.distinct("student_id", {**base, "status": "approved"}))
        else:
            paid_students = approved
        out.append({"type_id": tid, "type_name": t["name"], "approved": approved, "pending": pending,
                    "rejected": rejected, "recurring": recurring, "paid_students": paid_students,
                    "not_paid": max(total_students - paid_students, 0) if recurring
                    else max(total_students - approved - pending - rejected, 0)})
    return {"academic_year": year, "total_students": total_students, "stats": out}


class ManualPayment(BaseModel):
    student_id: str  # معرف الطالب (Mongo id)
    type_id: str
    other_label: Optional[str] = ""
    receipt_no: Optional[str] = ""
    amount: Optional[str] = ""
    receipt_date: Optional[str] = ""
    statement: Optional[str] = ""
    notes: Optional[str] = ""


@router.post("/fees/manual-payment")
async def manual_payment(data: ManualPayment, current_user: dict = Depends(get_current_user)):
    """✍️ تسجيل دفع يدوي: الموظف يعتبر الطالب دافعاً (دفع في الصندوق مثلاً) — معتمد فوراً"""
    if not can_manage_fees(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    if not ObjectId.is_valid(data.student_id):
        raise HTTPException(status_code=400, detail="معرف طالب غير صالح")
    student = await db.students.find_one({"_id": ObjectId(data.student_id)})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    if data.type_id == "other" and not (data.other_label or "").strip():
        raise HTTPException(status_code=400, detail="اكتب نوع الرسوم في الحقل الحر")
    year = await _academic_year(db)
    type_name, recurring = await _type_info(db, data.type_id, data.other_label)
    statement = (data.statement or "").strip()
    if recurring and not statement:
        raise HTTPException(status_code=400, detail=f"«{type_name}» رسوم متكررة — اكتب بيان الدفعة")
    sid = str(student["_id"])
    reviewer = current_user.get("full_name") or current_user.get("username", "")
    match: dict = {"student_id": sid, "type_id": data.type_id, "type_name": type_name, "academic_year": year}
    if recurring:
        match["statement"] = statement
    existing = await db.fee_receipts.find_one(match)
    if existing and existing.get("status") == "approved":
        raise HTTPException(status_code=400,
                            detail=f"دفعة «{statement}» معتمدة مسبقاً للطالب" if recurring else f"الطالب دافع مسبقاً لـ«{type_name}»")
    doc = {
        "student_id": sid, "type_id": data.type_id, "type_name": type_name,
        "academic_year": year, "image_base64": existing.get("image_base64", "") if existing else "",
        "receipt_no": (data.receipt_no or "").strip(), "amount": (data.amount or "").strip(),
        "receipt_date": (data.receipt_date or "").strip(),
        "statement": statement, "recurring": recurring,
        "notes": (data.notes or "").strip(), "status": "approved", "manual_entry": True,
        "rejection_reason": "", "uploaded_at": _now(), "reviewed_by": reviewer, "reviewed_at": _now(),
    }
    if existing:
        await db.fee_receipts.update_one({"_id": existing["_id"]}, {"$set": doc})
    else:
        await db.fee_receipts.insert_one(doc)
    _lbl = f"{type_name}" + (f" — {statement}" if statement else "")
    await _notify_student(db, student, f"✅ تم تسجيلك دافعاً — {_lbl}",
                          f"سجّلت الإدارة دفعك لرسوم «{_lbl}» للعام {year}. حالتك الآن: دافع.")
    await log_activity(current_user, "fee_manual_payment", "fee_receipt", sid, student.get("full_name", ""),
                       {"summary": f"تسجيل دفع يدوي «{_lbl}» للطالب {student.get('full_name', '؟')}" + (f" — سند {data.receipt_no}" if data.receipt_no else "")})
    return {"message": f"تم تسجيل {student.get('full_name', '')} دافعاً لـ«{_lbl}» ✅"}


@router.get("/fees/unpaid-export")
async def export_unpaid(type_id: str, current_user: dict = Depends(get_current_user)):
    """📄 تصدير Excel بغير الدافعين لنوع رسوم (لتسليمه لإدارة المالية)"""
    if not can_manage_fees(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    year = await _academic_year(db)
    type_name = await _type_name(db, type_id)
    receipts = {}
    approved_ids = set()
    async for r in db.fee_receipts.find({"type_id": type_id, "academic_year": year}, {"student_id": 1, "status": 1}):
        if r["status"] == "approved":
            approved_ids.add(r["student_id"])
        else:
            receipts[r["student_id"]] = r["status"]
    deps = {str(d["_id"]): d async for d in db.departments.find({}, {"name": 1, "faculty_id": 1})}
    facs = {str(f["_id"]): f.get("name", "") async for f in db.faculties.find({}, {"name": 1})}

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = "غير الدافعين"
    ws.sheet_view.rightToLeft = True
    ws.merge_cells("A1:G1")
    ws["A1"] = f"كشف غير الدافعين — {type_name} — العام الجامعي {year}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    headers = ["م", "رقم القيد", "اسم الطالب", "الكلية", "القسم", "المستوى", "حالة السند"]
    fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
    status_txt = {"pending": "قيد المراجعة", "rejected": "مرفوض — بانتظار إعادة الرفع"}
    row = 3
    n = 1
    async for s in db.students.find({"is_active": True}).sort([("department_id", 1), ("level", 1), ("full_name", 1)]):
        if str(s["_id"]) in approved_ids:
            continue
        st = receipts.get(str(s["_id"]))
        d = deps.get(s.get("department_id", ""), {})
        vals = [n, s.get("student_id", ""), s.get("full_name", ""),
                facs.get(str(d.get("faculty_id", "")), ""), d.get("name", ""),
                s.get("level", ""), status_txt.get(st, "لم يرفع السند")]
        for i, v in enumerate(vals, 1):
            ws.cell(row=row, column=i, value=v).alignment = Alignment(horizontal="center")
        row += 1
        n += 1
    for col, w in zip("ABCDEFG", [6, 14, 34, 22, 24, 10, 24]):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await log_activity(current_user, "fee_unpaid_export", "fee_receipt", "", None,
                       {"summary": f"تصدير كشف غير الدافعين لـ«{type_name}» ({n - 1} طالباً)"})
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=unpaid.xlsx"})


@router.post("/fees/receipts/{receipt_id}/unapprove")
async def unapprove_receipt(receipt_id: str, current_user: dict = Depends(get_current_user)):
    """↩️ التراجع عن اعتماد سند بالخطأ — يعود إلى «قيد المراجعة»"""
    if not can_manage_fees(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    r = await db.fee_receipts.find_one({"_id": ObjectId(receipt_id)})
    if not r:
        raise HTTPException(status_code=404, detail="السند غير موجود")
    if r.get("status") != "approved":
        raise HTTPException(status_code=400, detail="السند ليس معتمداً")
    await db.fee_receipts.update_one({"_id": r["_id"]}, {"$set": {
        "status": "pending", "reviewed_by": "", "reviewed_at": "", "rejection_reason": ""}})
    student = await db.students.find_one({"_id": ObjectId(r["student_id"])}) if ObjectId.is_valid(r["student_id"]) else None
    if student:
        await _notify_student(db, student, f"🔄 أُعيد سندك للمراجعة — {r['type_name']}",
                              f"أعادت الإدارة سند «{r['type_name']}»{(' — ' + r.get('statement')) if r.get('statement') else ''} إلى قيد المراجعة.")
    await log_activity(current_user, "fee_receipt_unapprove", "fee_receipt", receipt_id,
                       (student or {}).get("full_name", ""),
                       {"summary": f"التراجع عن اعتماد سند «{r['type_name']}» للطالب {(student or {}).get('full_name', '؟')}"})
    return {"message": "أُعيد السند إلى قيد المراجعة"}


class RenewalStatusBody(BaseModel):
    student_ids: list


@router.post("/fees/renewal-status")
async def renewal_status(data: RenewalStatusBody, current_user: dict = Depends(get_current_user)):
    """🟢 حالة تجديد القيد لمجموعة طلاب دفعة واحدة (لشارة جدول الطلاب)"""
    if current_user.get("role") == "student":
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    await _seed_types(db)
    t = await db.fee_types.find_one({"key": "renewal"})
    year = await _academic_year(db)
    ids = [str(s) for s in (data.student_ids or [])][:2000]
    out = {sid: "not_paid" for sid in ids}
    rank = {"approved": 3, "pending": 2, "rejected": 1}
    async for r in db.fee_receipts.find({"type_id": str(t["_id"]), "academic_year": year,
                                         "student_id": {"$in": ids}}, {"student_id": 1, "status": 1}):
        cur = out.get(r["student_id"], "not_paid")
        if rank.get(r["status"], 0) > rank.get(cur, 0):
            out[r["student_id"]] = r["status"]
    return {"statuses": out, "academic_year": year}


@router.get("/fees/students/{student_id}/receipts")
async def student_receipts(student_id: str, current_user: dict = Depends(get_current_user)):
    """💰 سجل سدادات طالب عبر كل الأعوام (للإدارة/المالية)"""
    if not can_manage_fees(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    receipts = await db.fee_receipts.find({"student_id": student_id}, {"image_base64": 0}).sort("uploaded_at", -1).to_list(500)
    return {"receipts": [{
        "id": str(r["_id"]), "type_name": r.get("type_name", ""), "statement": r.get("statement", ""),
        "academic_year": r.get("academic_year", ""), "receipt_no": r.get("receipt_no", ""),
        "amount": r.get("amount", ""), "receipt_date": r.get("receipt_date", ""),
        "date_warning": _date_warning(r.get("receipt_date", ""), r.get("academic_year", "")),
        "status": r.get("status", ""), "status_label": STATUS_LABELS.get(r.get("status"), ""),
        "rejection_reason": r.get("rejection_reason", ""), "manual_entry": bool(r.get("manual_entry")),
        "reviewed_by": r.get("reviewed_by", ""), "uploaded_at": r.get("uploaded_at", ""),
    } for r in receipts]}


@router.get("/fees/payment-report")
async def payment_report(date_from: str, date_to: str, student_ids: Optional[str] = None,
                         status: Optional[str] = "approved", fmt: str = "json",
                         current_user: dict = Depends(get_current_user)):
    """📊 تقرير سدادات طالب/عدة طلاب بين تاريخين (التاريخ الفعلي = تاريخ السند الورقي أو تاريخ الرفع)
    fmt: json | excel | pdf"""
    if not can_manage_fees(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    q: dict = {}
    if status and status != "all":
        q["status"] = status
    sid_list = [s.strip() for s in (student_ids or "").split(",") if s.strip()]
    if sid_list:
        q["student_id"] = {"$in": sid_list}
    receipts = await db.fee_receipts.find(q, {"image_base64": 0}).to_list(20000)
    rows = []
    for r in receipts:
        eff = (r.get("receipt_date") or "").strip() or (r.get("uploaded_at") or "")[:10]
        if not (date_from <= eff <= date_to):
            continue
        rows.append({**r, "_eff_date": eff})
    rows.sort(key=lambda x: x["_eff_date"])
    sids = [ObjectId(r["student_id"]) for r in rows if ObjectId.is_valid(r["student_id"])]
    smap = {str(s["_id"]): s for s in await db.students.find({"_id": {"$in": sids}}).to_list(20000)}
    deps = {str(d["_id"]): d.get("name", "") for d in await db.departments.find({}, {"name": 1}).to_list(500)}
    data = []
    total_amount = 0.0
    for r in rows:
        s = smap.get(r["student_id"], {})
        amt = r.get("amount", "")
        try:
            total_amount += float(str(amt).replace(",", "") or 0)
        except Exception:
            pass
        data.append({
            "id": str(r["_id"]), "date": r["_eff_date"], "student_name": s.get("full_name", "؟"),
            "student_number": s.get("student_id", ""), "department_name": deps.get(s.get("department_id", ""), ""),
            "type_name": r.get("type_name", ""), "statement": r.get("statement", ""),
            "receipt_no": r.get("receipt_no", ""), "amount": amt,
            "status_label": STATUS_LABELS.get(r.get("status"), ""),
            "manual_entry": bool(r.get("manual_entry")), "reviewed_by": r.get("reviewed_by", ""),
            "academic_year": r.get("academic_year", ""),
        })
    title = f"تقرير السدادات من {date_from} إلى {date_to}"
    if len(sid_list) == 1 and data:
        title += f" — الطالب: {data[0]['student_name']}"
    if fmt == "json":
        return {"title": title, "count": len(data), "total_amount": total_amount, "rows": data}

    headers = ["م", "التاريخ", "رقم القيد", "اسم الطالب", "القسم", "نوع الرسوم", "البيان", "رقم السند", "المبلغ", "الحالة", "عمّده"]
    table = [[i + 1, d["date"], d["student_number"], d["student_name"], d["department_name"],
              d["type_name"], d["statement"], d["receipt_no"], d["amount"],
              d["status_label"] + (" (يدوي)" if d["manual_entry"] else ""), d["reviewed_by"]]
             for i, d in enumerate(data)]

    import io
    from fastapi.responses import StreamingResponse
    if fmt == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "السدادات"
        ws.sheet_view.rightToLeft = True
        ws.merge_cells("A1:K1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].alignment = Alignment(horizontal="center")
        fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = fill
            c.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(table, 3):
            for ci, v in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=v).alignment = Alignment(horizontal="center")
        ws.cell(row=len(table) + 3, column=8, value="الإجمالي:").font = Font(bold=True)
        ws.cell(row=len(table) + 3, column=9, value=total_amount).font = Font(bold=True)
        for col, w in zip("ABCDEFGHIJK", [5, 12, 12, 30, 20, 20, 20, 12, 12, 16, 16]):
            ws.column_dimensions[col].width = w
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=payments.xlsx"})

    # PDF
    import arabic_reshaper
    from bidi.algorithm import get_display
    from pathlib import Path
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    def ar(t):
        return get_display(arabic_reshaper.reshape(str(t if t is not None else "")))

    font_path = Path(__file__).parent.parent / "fonts" / "Amiri-Regular.ttf"
    if "Amiri" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("Amiri", str(font_path)))
    bold_path = Path(__file__).parent.parent / "fonts" / "Amiri-Bold.ttf"
    if bold_path.exists() and "Amiri-Bold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("Amiri-Bold", str(bold_path)))
    BOLD = "Amiri-Bold" if "Amiri-Bold" in pdfmetrics.getRegisteredFontNames() else "Amiri"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=10 * mm, leftMargin=10 * mm,
                            topMargin=12 * mm, bottomMargin=10 * mm)
    elems = [Paragraph(ar(title), ParagraphStyle("t", fontName=BOLD, fontSize=14, alignment=1)), Spacer(1, 6 * mm)]
    pdf_rows = [[ar(h) for h in headers[::-1]]]
    for row in table:
        pdf_rows.append([ar(v) for v in row[::-1]])
    pdf_rows.append([ar("")] * 9 + [ar(f"{total_amount:,.0f}"), ar("الإجمالي")])
    tbl = Table(pdf_rows, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Amiri"),
        ("FONTNAME", (0, 0), (-1, 0), BOLD),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565C0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B0BEC5")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
    ]))
    elems.append(tbl)
    doc.build(elems)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=payments.pdf"})


class RemindBody(BaseModel):
    type_id: str


@router.post("/fees/remind-unpaid")
async def remind_unpaid(data: RemindBody, current_user: dict = Depends(get_current_user)):
    if not can_manage_fees(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    year = await _academic_year(db)
    type_name = await _type_name(db, data.type_id)
    covered = {r["student_id"] async for r in db.fee_receipts.find(
        {"type_id": data.type_id, "academic_year": year, "status": {"$in": ["approved", "pending"]}}, {"student_id": 1})}
    title = f"💰 تذكير: {type_name}"
    msg = f"لم يصلنا سند دفع «{type_name}» للعام {year}. يرجى رفع السند من التطبيق."
    sent = 0
    async for s in db.students.find({"is_active": True, "user_id": {"$exists": True, "$ne": ""}}):
        if str(s["_id"]) in covered:
            continue
        await _notify_student(db, s, title, msg)
        sent += 1
    await log_activity(current_user, "fee_remind_unpaid", "fee_receipt", "", None,
                       {"summary": f"تذكير جماعي بسند «{type_name}» — أُرسل لـ{sent} طالباً"})
    return {"message": f"أُرسل التذكير إلى {sent} طالباً غير دافع", "sent": sent}
