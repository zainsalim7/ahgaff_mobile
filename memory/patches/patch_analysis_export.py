import ast

P = "/app/backend/backend/routes/grades.py"
s = open(P, encoding="utf-8").read()

def rep(old, new):
    global s
    assert s.count(old) == 1, f"anchor ({s.count(old)}): {old[:70]!r}"
    s = s.replace(old, new)

# 1) استخراج منطق التحليل إلى دالة قابلة لإعادة الاستخدام
rep('''@router.get("/grades/analysis/{import_id}")
async def analyze_import(import_id: str, current_user: dict = Depends(get_current_user)):
    """تحليل متدرج لكشف مستورد: سعي فقط ← سعي+دور أول ← نتيجة مكتملة"""
    if not _can_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    imp = await db.grade_imports.find_one({"_id": ObjectId(import_id)})''',
'''async def _compute_analysis(db, import_id: str) -> dict:
    """تحليل متدرج لكشف مستورد: سعي فقط ← سعي+دور أول ← نتيجة مكتملة"""
    imp = await db.grade_imports.find_one({"_id": ObjectId(import_id)})''')

rep('''        "dismissed": dismissed,
    }


@router.get("/grades/analysis-cumulative")''',
'''        "dismissed": dismissed,
    }


@router.get("/grades/analysis/{import_id}")
async def analyze_import(import_id: str, current_user: dict = Depends(get_current_user)):
    if not _can_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    return await _compute_analysis(get_db(), import_id)


@router.get("/grades/analysis-cumulative")''')

# 2) نقاط التصدير في نهاية الملف
s += '''

# ===== 📤 تصدير تحليل النتيجة (PDF / Excel) لمجلس الكلية =====

_STAGE_LABELS = {"saai": "السعي فقط (إنذار مبكر)", "round1": "بعد الدور الأول (نتائج أولية)", "final": "النتيجة المكتملة"}


def _analysis_fname(a: dict, ext: str) -> str:
    info = a.get("info", {})
    parts = ["تحليل النتيجة"]
    if info.get("level"):
        parts.append(f"م{info['level']}")
    if info.get("semester_no"):
        parts.append(f"فصل {info['semester_no']}")
    if info.get("academic_year"):
        parts.append(str(info["academic_year"]))
    parts.append(datetime.now(timezone.utc).strftime("%Y-%m-%d"))
    return " - ".join(parts) + f".{ext}"


def _analysis_meta_rows(a: dict):
    info = a.get("info", {})
    rows = [["الكشف", info.get("filename", "")],
            ["المستوى", info.get("level", "")],
            ["الفصل", info.get("semester_no", "")],
            ["العام الجامعي", info.get("academic_year", "")],
            ["الدفعة", info.get("batch_no", "")],
            ["عدد الطلاب", info.get("students", 0)],
            ["مرحلة التحليل", _STAGE_LABELS.get(a.get("stage", ""), a.get("stage", ""))]]
    if a.get("stage") != "saai":
        sm = a.get("summary", {})
        rows += [["ناجح بكل المواد", sm.get("passed", 0)],
                 ["لديهم مواد راسبة (دور ثانٍ)", sm.get("second_round", 0)],
                 ["مفصولون", sm.get("dismissed", 0)]]
    return rows


def _analysis_course_table(a: dict):
    saai = a.get("stage") == "saai"
    if saai:
        head = ["المقرر", "الساعات", "العدد", "متوسط السعي", "الأعلى", "الأدنى", "سعي ضعيف", "غياب"]
        rows = [[c["course"], c.get("credits", ""), c["count"], c["avg"], c["max"], c["min"],
                 c.get("weak", 0), c.get("absent", 0)] for c in a.get("course_stats", [])]
    else:
        head = ["المقرر", "الساعات", "العدد", "المتوسط", "الأعلى", "الأدنى", "راسبون", "نسبة النجاح %", "متوسط السعي", "غياب"]
        rows = [[c["course"], c.get("credits", ""), c["count"], c["avg"], c["max"], c["min"],
                 c.get("fail_count", 0), c.get("pass_rate", ""),
                 c.get("avg_saai", "") if c.get("avg_saai") is not None else "", c.get("absent", 0)]
                for c in a.get("course_stats", [])]
    return head, rows


@router.get("/grades/analysis/{import_id}/export/excel")
async def export_analysis_excel(import_id: str, current_user: dict = Depends(get_current_user)):
    if not _can_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    a = await _compute_analysis(get_db(), import_id)
    from urllib.parse import quote
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet(title, head, rows):
        ws = wb.create_sheet(title=title[:31])
        ws.sheet_view.rightToLeft = True
        ws.append([f"تقرير تحليل النتيجة — {title}"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(head), 2))
        t = ws.cell(row=1, column=1)
        t.font = Font(bold=True, size=13, color="1A237E")
        t.alignment = Alignment(horizontal="center")
        ws.append(head)
        for c in ws[2]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="3949AB")
            c.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append(r)
        for col in ws.iter_cols(min_row=2):
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(w + 4, 10), 50)
        return ws

    add_sheet("الملخص", ["البيان", "القيمة"], _analysis_meta_rows(a))
    head, rows = _analysis_course_table(a)
    add_sheet("إحصاءات المقررات", head, rows)
    if a.get("top"):
        add_sheet("الأوائل", ["م", "الاسم", "رقم القيد", "المعدل"],
                  [[i + 1, t.get("name", ""), t.get("reg_no", ""), t.get("avg", "")] for i, t in enumerate(a["top"])])
    if a.get("second_round"):
        add_sheet("الدور الثاني", ["الاسم", "رقم القيد", "عدد المواد", "المواد الراسب فيها"],
                  [[x.get("name", ""), x.get("reg_no", ""), len(x.get("courses", [])), "، ".join(x.get("courses", []))] for x in a["second_round"]])
    if a.get("fail_lists") and any(v for v in a["fail_lists"].values()):
        add_sheet("الراسبون بالمقرر", ["المقرر", "العدد", "الأسماء"],
                  [[cn, len(names), "، ".join(names)] for cn, names in a["fail_lists"].items() if names])
    if a.get("saai_alerts"):
        add_sheet("إنذار السعي", ["الاسم", "رقم القيد", "المقررات (سعي < النصف)"],
                  [[x.get("name", ""), x.get("reg_no", ""), "، ".join(x.get("courses", []))] for x in a["saai_alerts"]])
    if a.get("dismissed"):
        add_sheet("المفصولون", ["الاسم", "الملاحظة"],
                  [[x.get("name", ""), x.get("note", "")] for x in a["dismissed"]])

    out = io.BytesIO()
    wb.save(out)
    fname = quote(_analysis_fname(a, "xlsx"))
    return StreamingResponse(io.BytesIO(out.getvalue()),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}",
                                      "X-Filename": fname})


@router.get("/grades/analysis/{import_id}/export/pdf")
async def export_analysis_pdf(import_id: str, current_user: dict = Depends(get_current_user)):
    if not _can_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    a = await _compute_analysis(get_db(), import_id)
    from urllib.parse import quote
    import arabic_reshaper
    from bidi.algorithm import get_display
    from pathlib import Path
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER

    def ar(t):
        return get_display(arabic_reshaper.reshape(str(t if t is not None else "")))

    fonts_dir = Path(__file__).parent.parent / "fonts"
    if "Amiri" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("Amiri", str(fonts_dir / "Amiri-Regular.ttf")))
    if (fonts_dir / "Amiri-Bold.ttf").exists() and "Amiri-Bold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("Amiri-Bold", str(fonts_dir / "Amiri-Bold.ttf")))
    BOLD = "Amiri-Bold" if "Amiri-Bold" in pdfmetrics.getRegisteredFontNames() else "Amiri"

    title_st = ParagraphStyle("t", fontName=BOLD, fontSize=16, alignment=TA_CENTER, spaceAfter=4)
    sub_st = ParagraphStyle("s", fontName="Amiri", fontSize=11, alignment=TA_CENTER, textColor=colors.HexColor("#455a64"))
    sec_st = ParagraphStyle("h", fontName=BOLD, fontSize=12.5, alignment=TA_RIGHT, textColor=colors.HexColor("#1a237e"), spaceBefore=10, spaceAfter=4)

    def rtl_table(head, rows, col_widths=None, highlight_col=None):
        data = [[ar(h) for h in reversed(head)]] + [[ar(v) for v in reversed(r)] for r in rows]
        t = Table(data, colWidths=list(reversed(col_widths)) if col_widths else None, repeatRows=1)
        style = [("FONTNAME", (0, 0), (-1, -1), "Amiri"),
                 ("FONTNAME", (0, 0), (-1, 0), BOLD),
                 ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                 ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3949ab")),
                 ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                 ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                 ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                 ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#b0bec5")),
                 ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f7ff")]),
                 ("TOPPADDING", (0, 0), (-1, -1), 3),
                 ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]
        t.setStyle(TableStyle(style))
        return t

    info = a.get("info", {})
    story = [
        Paragraph(ar("جامعة الأحقاف"), title_st),
        Paragraph(ar("تقرير تحليل النتيجة"), title_st),
        Paragraph(ar(f"{info.get('filename', '')} — المستوى {info.get('level', '')} · الفصل {info.get('semester_no', '')} · {info.get('academic_year', '')}"), sub_st),
        Paragraph(ar(f"مرحلة التحليل: {_STAGE_LABELS.get(a.get('stage', ''), '')} — عدد الطلاب: {info.get('students', 0)}"), sub_st),
        Spacer(1, 6 * mm),
    ]

    if a.get("stage") != "saai":
        sm = a.get("summary", {})
        story.append(rtl_table(["ناجح بكل المواد", "لديهم مواد راسبة (دور ثانٍ)", "مفصولون"],
                               [[sm.get("passed", 0), sm.get("second_round", 0), sm.get("dismissed", 0)]],
                               col_widths=[85 * mm, 85 * mm, 85 * mm]))
        story.append(Spacer(1, 4 * mm))

    story.append(Paragraph(ar("📚 إحصاءات المقررات"), sec_st))
    head, rows = _analysis_course_table(a)
    story.append(rtl_table(head, rows))

    if a.get("top"):
        story.append(Paragraph(ar("🏆 العشرة الأوائل"), sec_st))
        story.append(rtl_table(["م", "الاسم", "رقم القيد", "المعدل"],
                               [[i + 1, t.get("name", ""), t.get("reg_no", ""), t.get("avg", "")] for i, t in enumerate(a["top"])]))

    if a.get("second_round"):
        story.append(Paragraph(ar(f"📘 طلاب الدور الثاني ({len(a['second_round'])})"), sec_st))
        story.append(rtl_table(["الاسم", "رقم القيد", "المواد الراسب فيها"],
                               [[x.get("name", ""), x.get("reg_no", ""), "، ".join(x.get("courses", []))] for x in a["second_round"]]))

    if a.get("saai_alerts"):
        story.append(Paragraph(ar(f"⚠️ إنذار مبكر — سعي أقل من النصف ({len(a['saai_alerts'])})"), sec_st))
        story.append(rtl_table(["الاسم", "رقم القيد", "المقررات"],
                               [[x.get("name", ""), x.get("reg_no", ""), "، ".join(x.get("courses", []))] for x in a["saai_alerts"]]))

    if a.get("dismissed"):
        story.append(Paragraph(ar(f"⛔ المفصولون ({len(a['dismissed'])})"), sec_st))
        story.append(rtl_table(["الاسم", "الملاحظة"],
                               [[x.get("name", ""), x.get("note", "")] for x in a["dismissed"]]))

    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph(ar(f"أُصدر آلياً من نظام إدارة الجامعة — {datetime.now(timezone.utc).strftime('%Y-%m-%d')}"),
                           ParagraphStyle("f", fontName="Amiri", fontSize=8.5, alignment=TA_CENTER, textColor=colors.HexColor("#90a4ae"))))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    doc.build(story)
    fname = quote(_analysis_fname(a, "pdf"))
    return StreamingResponse(io.BytesIO(buf.getvalue()), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}",
                                      "X-Filename": fname})
'''

ast.parse(s)
open(P, "w", encoding="utf-8").write(s)
print("grades.py OK")
